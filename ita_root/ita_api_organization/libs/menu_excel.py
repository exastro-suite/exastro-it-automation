#   Copyright 2022 NEC Corporation
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.

from concurrent.futures import process
from curses import keyname
from dataclasses import dataclass
from inspect import Parameter
from zipfile import BadZipfile
import openpyxl
import base64
from pyrsistent import m

import pytz
import six
import datetime
import json

from common_libs.common import *  # noqa: F403
from flask import jsonify, session, g
from copy import copy
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import absolute_coordinate
from common_libs.loadtable import *
from libs import menu_maintenance_all, menu_info

def collect_excel_all(objdbca, organization_id, workspace_id, menu, menu_record, menu_table_link_record):
    """
        全件のExcelを取得する

        ARGS:
            objdbca: DB接クラス  DBConnectWs()
            organization_id: Organization ID
            workspace_id: Workspace ID
            menu: メニュー名
            menu_record: メニュー管理のレコード
            menu_table_link_record: メニュー-テーブル紐付管理のレコード
        RETURN:
            Excelをbase64した値
    """
    # 言語設定取得
    lang = g.LANGUAGE
    
    # make storage directory for excel
    strage_path = os.environ.get('STORAGEPATH')
    excel_dir = strage_path + "/".join([organization_id, workspace_id]) + "/tmp/excel"
    if not os.path.isdir(excel_dir):
        os.makedirs(excel_dir)
        g.applogger.debug("made excel_dir")
    
    # テーブル名
    t_common_column_group = 'T_COMN_COLUMN_GROUP'
    t_common_menu_column_link = 'T_COMN_MENU_COLUMN_LINK'
    
    # メニューのカラム情報を取得
    objmenu = load_table.loadTable(objdbca, menu)   # noqa: F405
    mode = 'excel'
    result_data = objmenu.rest_filter({}, mode)
    
    #### result_code,msg未対応
    result = {
        "result": "result_code", #result_data[0],
        "data": result_data, #result_data[1],
        "message": "msg" #result_data[2]
    }
    
    # 対象メニューを特定するためのIDを取得する
    menu_id = menu_record[0].get('MENU_ID')
    menu_name = menu_record[0].get('MENU_NAME_' + lang.upper())
    file_name = menu_name + '_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
    file_path = excel_dir + '/' + file_name
    
    # 『メニュー-カラム紐付管理』テーブルから対象のデータを取得
    retList_t_common_menu_column_link = objdbca.table_select(t_common_menu_column_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s ORDER BY COLUMN_DISP_SEQ ASC', [menu_id, 0])
    if not retList_t_common_menu_column_link:
        log_msg_args = [menu]
        api_msg_args = [menu]
        raise AppException("499-00005", log_msg_args, api_msg_args)  # noqa: F405
    
    # 使用するCOL_GROUP_IDだけを配列に確保しておく
    active_col_group_id = ['']
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        tmp = dict_menu_column.get('COL_GROUP_ID')
        if tmp is not None and tmp not in active_col_group_id:
            active_col_group_id.append(tmp)
    
    # IDColumn項目のプルダウン一覧の取得
    pulldown_list = menu_info.collect_pulldown_list(objdbca, menu, menu_record)
    
    # 登録更新廃止復活フラグを取得する
    menu_table_link_list = []
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_INSERT_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_UPDATE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_DISUSE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_REUSE_FLAG'))
    
    # 色の設定
    # 背景色
    fill_wh = PatternFill(fill_type='solid', fgColor='FFFFFF')
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_lt = Alignment(horizontal='left', vertical='top')
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    sideDash = Side(border_style="dashed", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    borderDash = Border(left=side, right=side, top=sideDash, bottom=sideDash)
    
    # ワークブックの新規作成と保存
    wb = Workbook()
    
    # マスタ
    msg = g.appmsg.get_api_message('MSG-30012')
    wb.create_sheet(msg)
    # フィルタ条件
    msg = g.appmsg.get_api_message('MSG-30013')
    wb.create_sheet(msg)
    
    # 「マスタ」シート作成
    # 「マスタ」シートの名前の定義をメインのシートの入力規則で使用するため「マスタ」シートから作成する
    name_define_list = make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, menu_table_link_list, pulldown_list)
    
    # シートを指定して編集
    ws = wb.active
    ws.title = menu_name[:31]
    
    # ヘッダー部編集
    startRow = 1
    startClm = 4
    
    # 『カラムグループ管理』テーブルからカラムグループ一覧を取得
    ret = objdbca.table_select(t_common_column_group, 'WHERE COL_GROUP_ID IN %s AND DISUSE_FLAG = %s', [active_col_group_id, 0])
    dict_column_group_id = {}
    dict_column_group_id_name = {}
    if ret:
        for recode in ret:
            dict_column_group_id[recode.get('COL_GROUP_ID')] = recode.get('PA_COL_GROUP_ID')
            # IDと名前の辞書も作成する
            dict_column_group_id_name[recode.get('COL_GROUP_ID')] = recode.get('COL_GROUP_NAME_' + lang.upper())
    
    # ヘッダーの階層の深さを調べる
    depth = get_col_group_depth(retList_t_common_menu_column_link, dict_column_group_id)
    
    # エクセルに表示するヘッダー項目を二次元配列に構築する
    excel_header_list, header_order = create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id, menu_table_link_list)
    
    # 1行目（項目名）のヘッダーを作成する
    ws = create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border)
    
    startRow = startRow + depth
    startDetailRow = startRow + 7
    
    # 2行目以降を作成する
    # 固定部分
    ws = make_template(ws, startRow, font_bl, fill_gr, borderDash, depth)
    
    # エクセルヘッダー部のカラム情報を作成する
    ws, dataVaridationDict = create_column_info(lang, ws, startRow, startClm, retList_t_common_menu_column_link, name_define_list, menu_table_link_list)
    
    # 明細部編集
    # 明細1行目編集
    ws, dataVaridationDict = detail_first_line_format(ws, startRow, dataVaridationDict)
    
    # parameterとfileのリストを取得
    mylist = result.get("data")[1]
    
    # ウィンドウ枠の固定
    ws.freeze_panes = ws.cell(row=startRow + 7, column=4).coordinate
    
    for param in mylist:
        for k, v in param.items():
            if k == 'parameter':
                for i, (key, value) in enumerate(v.items()):
                    if i == 0:
                        # 初回で固定部分の設定をしておく
                        ws.cell(row=startRow + 7, column=1).fill = fill_wh
                        ws.cell(row=startRow + 7, column=1, value='')
                        
                        ws.cell(row=startRow + 7, column=2).fill = fill_wh
                        ws.cell(row=startRow + 7, column=2, value='')
                        
                        # 実行処理種別
                        ws.cell(row=startRow + 7, column=3).font = font_bl
                        ws.cell(row=startRow + 7, column=3).alignment = al_cc
                        ws.cell(row=startRow + 7, column=3).border = border
                        ws.cell(row=startRow + 7, column=3, value='-')
                        # 入力規則の設定
                        dv = DataValidation(type='list', formula1='FILTER_ROW_EDIT_BY_FILE')
                        dv.add(ws.cell(row=startRow + 7, column=3))
                        ws.add_data_validation(dv)
                        dataVaridationDict[get_column_letter(3)] = 'FILTER_ROW_EDIT_BY_FILE'
                    
                    if key == 'discard':
                        # 廃止
                        msg = g.appmsg.get_api_message('MSG-30006')
                        if value == '1':
                            value = msg
                        else:
                            value = ''
                    
                    # dict or list 形式を json.dumps
                    if isinstance(value, dict) or isinstance(value, list):
                        value = json.dumps(value, ensure_ascii=False)

                    if key in header_order:
                        column_num = header_order.index(key) + 4
                    else:
                        continue
                    
                    ws.cell(row=startRow + 7, column=column_num).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                    ws.cell(row=startRow + 7, column=column_num).font = font_bl
                    ws.cell(row=startRow + 7, column=column_num).border = border
                    ws.cell(row=startRow + 7, column=column_num, value=value)
                    if key in name_define_list:
                        dv = DataValidation(type='list', formula1=key)
                        dv.add(ws.cell(row=startRow + 7, column=column_num))
                        ws.add_data_validation(dv)
                        if key not in dataVaridationDict:
                            dataVaridationDict[get_column_letter(column_num)] = key
                startRow = startRow + 1
    
    # 空行追加処理
    ws = create_blank_line(ws, dataVaridationDict, 10)
    
    # 登録が×の列をグレーにする
    # 明細の数を求める
    column_num = ws.max_column + 1
    row_num = ws.max_row - startDetailRow + 1
    for col_i in range(4, column_num):
        for row_j in range(row_num):
            if ws.cell(row=depth + 1, column=col_i).value == '×':
                # セルの背景色をグレーにする
                ws.cell(row=startDetailRow + row_j, column=col_i).fill = fill_gr
    
    # フッターを作成する
    ws = create_footer(ws, font_wh, al_lt, al_cc, fill_bl)
    
    # フィルタ条件シートを指定して編集
    # フィルタ条件
    msg = g.appmsg.get_api_message('MSG-30013')
    ws_filter = wb[msg]
    # 出力日時
    msg = g.appmsg.get_api_message('MSG-30014')
    ws_filter['A1'] = msg
    # 廃止
    msg = g.appmsg.get_api_message('MSG-30006')
    ws_filter['D1'] = msg
    ws_filter['A2'] = datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
    
    # 廃止含まず
    msg = g.appmsg.get_api_message('MSG-30016')
    ws_filter['D2'] = msg
    
    wb.save(file_path)  # noqa: E303
    
    # 編集してきたエクセルファイルをエンコードする
    wbEncode = file_encode(file_path)  # noqa: F405 F841
    # エンコード後wbは削除する
    os.remove(file_path)
    
    return wbEncode


# 「マスタ」シートを作成する
def make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, menu_table_link_list, pulldown_list):  # noqa: E302
    # マスタシートを指定して編集
    msg = g.appmsg.get_api_message('MSG-30012')
    ws_master = wb[msg]
    
    # 開始位置
    startRow_master = 1
    startClm_master = 6
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws_master.cell(row=1, column=3, value=msg)
    if menu_table_link_list[0] == '1':
        # 登録
        msg = g.appmsg.get_api_message('MSG-30004')
        ws_master.cell(row=2, column=3, value=msg)
    if menu_table_link_list[1] == '1':
        # 更新
        msg = g.appmsg.get_api_message('MSG-30005')
        ws_master.cell(row=3, column=3, value=msg)
    if menu_table_link_list[2] == '1':
        # 廃止
        msg = g.appmsg.get_api_message('MSG-30006')
        ws_master.cell(row=4, column=3, value=msg)
    if menu_table_link_list[3] == '1':
        # 復活
        msg = g.appmsg.get_api_message('MSG-30007')
        ws_master.cell(row=5, column=3, value=msg)
    
    # 空白行を消して詰める
    for i in reversed(range(1, 5)):
        if ws_master.cell(row=i, column=3).value == None:
            ws_master.delete_rows(i)
    
    # 名前の定義に使用する範囲を求める
    cnt = menu_table_link_list.count('1')
    
    if cnt >= 1:
        msg = g.appmsg.get_api_message('MSG-30012')
        new_range = DefinedName(name='FILTER_ROW_EDIT_BY_FILE', attr_text=msg + '!$C$2:$C$' + str(cnt + 1))
        wb.defined_names.append(new_range)
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws_master.cell(row=1, column=4, value=msg)
    
    # 廃止
    msg = g.appmsg.get_api_message('MSG-30006')
    ws_master.cell(row=1, column=5, value=msg)
    
    name_define_list = []
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        ws_master.cell(row=startRow_master, column=startClm_master + i, value=dict_menu_column.get('COLUMN_NAME_' + lang.upper()))
        column_name_rest = dict_menu_column.get('COLUMN_NAME_REST')
        if column_name_rest in pulldown_list:
            # 名前の定義開始、終了位置
            startCell = None
            endCell = None
            for j, value in enumerate(pulldown_list[column_name_rest].values(), 1):
                last_loop = len(pulldown_list[column_name_rest])
                ws_master.cell(row=startRow_master + j, column=startClm_master + i, value=value)
                
                if j == 1:
                    startCell = ws_master.cell(row=startRow_master + j, column=startClm_master + i).coordinate
                
                # 最後の要素で名前の定義を設定する
                if j == last_loop:
                    if startCell is not None:
                        # マスタ
                        msg = g.appmsg.get_api_message('MSG-30012')
                        endCell = ws_master.cell(row=startRow_master + j, column=startClm_master + i).coordinate
                        new_range = DefinedName(name=column_name_rest, attr_text=msg + '!' + absolute_coordinate(startCell) + ':' + absolute_coordinate(endCell))
                        wb.defined_names.append(new_range)
                        name_define_list.append(column_name_rest)
    
    return name_define_list


# エクセル固定部分の作成
def make_template(ws, startRow, font, fill, borderDash, depth):
    # フォント
    font_or = openpyxl.styles.Font(name='メイリオ', size=8, color='E65000', bold=True)
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # 背景色
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_ltw = Alignment(horizontal='left', vertical='top', wrapText=True)
    al_rt = Alignment(horizontal='right', vertical='top', wrapText=True)
    
    # 罫線
    side = Side(border_style="thin", color="000000")
    sideDash = Side(border_style="dashed", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    border_r = Border(left=side, right=None, top=side, bottom=side)
    border_l = Border(left=None, right=side, top=side, bottom=side)
    borderDash_r = Border(left=side, right=sideDash, top=side, bottom=side)
    borderDash_ltb = Border(left=sideDash, right=side, top=sideDash, bottom=sideDash)
    
    # 固定部分
    # 注意事項
    msg = g.appmsg.get_api_message('MSG-30001')
    ws.cell(row=1, column=1).font = font_wh
    ws.cell(row=1, column=1).alignment = al_cc
    ws.cell(row=1, column=1).fill = fill_bl
    ws.cell(row=1, column=1).border = border_r
    ws.cell(row=1, column=1, value=msg)
    
    ws.cell(row=1, column=2).font = font_wh
    ws.cell(row=1, column=2).alignment = al_cc
    ws.cell(row=1, column=2).fill = fill_bl
    ws.cell(row=1, column=2).border = border_l
    ws.cell(row=1, column=2, value='')
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws.cell(row=1, column=3).font = font_wh
    ws.cell(row=1, column=3).alignment = al_cc
    ws.cell(row=1, column=3).fill = fill_bl
    ws.cell(row=1, column=3).border = border
    ws.cell(row=1, column=3, value=msg)
    
    # 固定部分の結合
    alphaOrd = ord('A')
    for i in range(4):
        alphaChr = chr(alphaOrd)
        ws.merge_cells(alphaChr + '1:' + alphaChr + str(depth))
        alphaOrd = alphaOrd + 1
    
    msg = g.appmsg.get_api_message('MSG-30003')
    ws.cell(row=startRow, column=1).font = font
    ws.cell(row=startRow, column=1).fill = fill
    ws.cell(row=startRow, column=1).alignment = al_ltw
    ws.cell(row=startRow, column=1).border = borderDash_r
    ws.cell(row=startRow, column=1, value=msg)
    ws.merge_cells('A' + str(startRow) + ':' + 'A' + str(startRow + 3))
    
    # 登録更新廃止復活
    msg = g.appmsg.get_api_message('MSG-30004')
    ws.cell(row=startRow, column=2).font = font
    ws.cell(row=startRow, column=2).fill = fill
    ws.cell(row=startRow, column=2).alignment = al_cc
    ws.cell(row=startRow, column=2).border = borderDash_ltb
    ws.cell(row=startRow, column=2, value=msg)
    
    msg = g.appmsg.get_api_message('MSG-30005')
    ws.cell(row=startRow + 1, column=2).font = font
    ws.cell(row=startRow + 1, column=2).fill = fill
    ws.cell(row=startRow + 1, column=2).alignment = al_cc
    ws.cell(row=startRow + 1, column=2).border = borderDash_ltb
    ws.cell(row=startRow + 1, column=2, value=msg)
    
    msg = g.appmsg.get_api_message('MSG-30006')
    ws.cell(row=startRow + 2, column=2).font = font
    ws.cell(row=startRow + 2, column=2).fill = fill
    ws.cell(row=startRow + 2, column=2).alignment = al_cc
    ws.cell(row=startRow + 2, column=2).border = borderDash_ltb
    ws.cell(row=startRow + 2, column=2, value=msg)
    
    msg = g.appmsg.get_api_message('MSG-30007')
    ws.cell(row=startRow + 3, column=2).font = font
    ws.cell(row=startRow + 3, column=2).fill = fill
    ws.cell(row=startRow + 3, column=2).alignment = al_cc
    ws.cell(row=startRow + 3, column=2).border = borderDash_ltb
    ws.cell(row=startRow + 3, column=2, value=msg)
    
    # 実行処理種別
    for i in range(4):
        ws.cell(row=startRow + i, column=3).font = font
        ws.cell(row=startRow + i, column=3).fill = fill
        ws.cell(row=startRow + i, column=3).alignment = al_cc
        ws.cell(row=startRow + i, column=3).border = borderDash
        ws.cell(row=startRow + i, column=3, value='')
    
    # 改行
    msg = g.appmsg.get_api_message('MSG-30008')
    ws.cell(row=startRow + 4, column=1).font = font
    ws.cell(row=startRow + 4, column=1).fill = fill
    ws.cell(row=startRow + 4, column=1).alignment = al_ltw
    ws.cell(row=startRow + 4, column=1).border = border
    ws.cell(row=startRow + 4, column=1, value=msg)
    ws.merge_cells('A' + str(startRow + 4) + ':' + 'B' + str(startRow + 4))
    
    ws.cell(row=startRow + 4, column=3).font = font
    ws.cell(row=startRow + 4, column=3).fill = fill
    ws.cell(row=startRow + 4, column=3).alignment = al_ltw
    ws.cell(row=startRow + 4, column=3).border = border
    ws.cell(row=startRow + 4, column=3, value='')
    
    msg = g.appmsg.get_api_message('MSG-30009')
    ws.cell(row=startRow + 5, column=1).font = font
    ws.cell(row=startRow + 5, column=1).fill = fill
    ws.cell(row=startRow + 5, column=1).alignment = al_rt
    ws.cell(row=startRow + 5, column=1).border = border
    ws.cell(row=startRow + 5, column=1, value=msg)
    ws.merge_cells('A' + str(startRow + 5) + ':' + 'B' + str(startRow + 5))
    
    msg = g.appmsg.get_api_message('MSG-30010')
    ws.cell(row=startRow + 5, column=3).font = font_or
    ws.cell(row=startRow + 5, column=3).fill = fill
    ws.cell(row=startRow + 5, column=3).alignment = al_ltw
    ws.cell(row=startRow + 5, column=3).border = border
    ws.cell(row=startRow + 5, column=3, value=msg)
    
    ws.cell(row=startRow + 6, column=1).font = font_wh
    ws.cell(row=startRow + 6, column=1).fill = fill_bl
    ws.cell(row=startRow + 6, column=1).alignment = al_cc
    ws.cell(row=startRow + 6, column=1).border = border_r
    ws.cell(row=startRow + 6, column=1, value='')
    
    ws.cell(row=startRow + 6, column=2).font = font_wh
    ws.cell(row=startRow + 6, column=2).fill = fill_bl
    ws.cell(row=startRow + 6, column=2).alignment = al_cc
    ws.cell(row=startRow + 6, column=2).border = border_l
    ws.cell(row=startRow + 6, column=2, value='')
    
    # 実行処理種別
    msg = g.appmsg.get_api_message('MSG-30002')
    ws.cell(row=startRow + 6, column=3).font = font_wh
    ws.cell(row=startRow + 6, column=3).fill = fill_bl
    ws.cell(row=startRow + 6, column=3).alignment = al_cc
    ws.cell(row=startRow + 6, column=3).border = border
    ws.cell(row=startRow + 6, column=3, value=msg)
    
    return ws


# エクセル固定部分の作成(変更履歴ver)
def make_template_trace_history(ws, depth):
    # 文字色
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    
    # 背景色
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # 罫線
    side = Side(border_style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # 履歴通番
    msg = g.appmsg.get_api_message('MSG-30020')
    ws.cell(row=1, column=1).font = font_wh
    ws.cell(row=1, column=1).alignment = al_cc
    ws.cell(row=1, column=1).fill = fill_bl
    ws.cell(row=1, column=1).border = border
    ws.cell(row=1, column=1, value=msg)
    
    ws.cell(row=depth + 2, column=1).font = font_wh
    ws.cell(row=depth + 2, column=1).alignment = al_cc
    ws.cell(row=depth + 2, column=1).fill = fill_bl
    ws.cell(row=depth + 2, column=1).border = border
    ws.cell(row=depth + 2, column=1, value=msg)
    
    # 変更日時
    msg = g.appmsg.get_api_message('MSG-30021')
    ws.cell(row=1, column=2).font = font_wh
    ws.cell(row=1, column=2).alignment = al_cc
    ws.cell(row=1, column=2).fill = fill_bl
    ws.cell(row=1, column=2).border = border
    ws.cell(row=1, column=2, value=msg)
    
    ws.cell(row=depth + 2, column=2).font = font_wh
    ws.cell(row=depth + 2, column=2).alignment = al_cc
    ws.cell(row=depth + 2, column=2).fill = fill_bl
    ws.cell(row=depth + 2, column=2).border = border
    ws.cell(row=depth + 2, column=2, value=msg)
    # 変更日時はマイクロ秒部分を見えないように幅を設定する
    ws.column_dimensions[get_column_letter(2)].width = 15.7
    
    # 固定部分の結合
    alphaOrd = ord('A')
    for i in range(4):
        alphaChr = chr(alphaOrd)
        ws.merge_cells(alphaChr + '1:' + alphaChr + str(depth))
        alphaOrd = alphaOrd + 1
    
    # 空白
    ws.cell(row=depth + 1, column=1).font = font_wh
    ws.cell(row=depth + 1, column=1).fill = fill_gr
    ws.cell(row=depth + 1, column=1).alignment = al_cc
    ws.cell(row=depth + 1, column=1).border = border
    ws.cell(row=depth + 1, column=1, value='')
    
    ws.cell(row=depth + 1, column=2).font = font_wh
    ws.cell(row=depth + 1, column=2).fill = fill_gr
    ws.cell(row=depth + 1, column=2).alignment = al_cc
    ws.cell(row=depth + 1, column=2).border = border
    ws.cell(row=depth + 1, column=2, value='')
    
    return ws


def get_col_group_depth(list_menu_column_link, dict_col_group):  # noqa: E302
    depth = 1
    for dict_menu_column in list_menu_column_link:
        cnt = 1
        col_group_id = dict_menu_column.get('COL_GROUP_ID')
        if col_group_id is not None:
            # 親が無くなるまで繰り返す
            pa_search_flag = True
            while pa_search_flag:
                cnt = cnt + 1
                col_group_id = dict_col_group.get(col_group_id)
                if col_group_id is None:
                    pa_search_flag = False
                    if depth < cnt:
                        depth = cnt
    return depth


def recursive_get_pa_col_group_id(n, id, dict_group_id):
    if n < 1:
        return id
    
    pa_id = dict_group_id.get(id)
    
    return recursive_get_pa_col_group_id(n - 1, pa_id, dict_group_id)  # noqa: E303


# エクセルに表示するヘッダー項目を二次元配列に構築する
def create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id, menu_table_link_list):
    # 行を階層分追加する
    if depth > 1:
        ws.insert_rows(1, depth - 1)
    
    excel_header_list = [[] for i in range(depth)]
    header_order = []
    
    # 表示する項目を二次元配列に構築する
    for i in range(depth):
        for j, dict_menu_column in enumerate(retList_t_common_menu_column_link):
            column_name = dict_menu_column.get('COLUMN_NAME_' + lang.upper())
            column_name_rest = dict_menu_column.get('COLUMN_NAME_REST')
            input_item = dict_menu_column.get('INPUT_ITEM')
            view_item = dict_menu_column.get('VIEW_ITEM')
            column_class = dict_menu_column.get('COLUMN_CLASS')

            if input_item == '2' and view_item == '0':
                # excelに表示しない
                continue
            
            if column_class == '23':
                # excelに表示しない
                continue

            # 廃止フラグ
            msg = g.appmsg.get_api_message('MSG-30015')
            if column_name == msg:
                excel_header_list[depth - 1 - i].insert(0, column_name)
                if column_name_rest not in header_order:
                    header_order.insert(0, column_name_rest)
                continue
            
            if i == 0:
                excel_header_list[depth - 1 - i].append(column_name)
                if column_name_rest not in header_order:
                    header_order.append(column_name_rest)
            elif i == 1:
                group_id = dict_menu_column.get('COL_GROUP_ID')
                if group_id is None:
                    excel_header_list[depth - 1 - i].append(column_name)
                else:
                    column_group_name = dict_column_group_id_name.get(group_id)
                    if column_group_name is None:
                        column_group_name = ''
                    excel_header_list[depth - 1 - i].append(column_group_name)
            else:
                group_id = recursive_get_pa_col_group_id(i - 1, dict_menu_column.get('COL_GROUP_ID'), dict_column_group_id)
                if group_id is None:
                    excel_header_list[depth - 1 - i].append(column_name)
                else:
                    column_group_name = dict_column_group_id_name.get(group_id)
                    if column_group_name is None:
                        column_group_name = ''
                    excel_header_list[depth - 1 - i].append(column_group_name)
    
    # 親が一番上にくるようにリストを整える
    for i in range(len(excel_header_list[0])):
        for j in range(depth):
            if excel_header_list[j][i] == excel_header_list[depth - 1][i]:
                loopflg = True
                cnt = 1
                # 子が上にある場合、親と値を入れ替える
                while loopflg:
                    if j + cnt >= depth:
                        loopflg = False
                    elif excel_header_list[j][i] != excel_header_list[j + cnt][i]:
                        loopflg = False
                        excel_header_list[j][i] = excel_header_list[j + cnt][i]
                        excel_header_list[j + cnt][i] = excel_header_list[depth - 1][i]
                    cnt += 1
    
    return excel_header_list, header_order


# 1行目（項目名）のヘッダーを作成し、結合する
def create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border):
    # 1行目（項目名）のヘッダーを作成する
    for i, row in enumerate(excel_header_list):
        # 結合開始位置
        startCell = None
        # 前回ループ時の値との比較用
        tmp = None
        for j, data in enumerate(row):
            msg = g.appmsg.get_api_message('MSG-30015')
            if data == msg:
                # 廃止フラグだったら先頭(startClm行)に出力する
                ws.column_dimensions[ws.cell(row=startRow + i, column=startClm).column_letter].width = 2
                ws.cell(row=startRow + i, column=startClm).font = font_wh
                ws.cell(row=startRow + i, column=startClm).alignment = al_cc
                ws.cell(row=startRow + i, column=startClm).fill = fill_bl
                ws.cell(row=startRow + i, column=startClm).border = border
                ws.cell(row=startRow + i, column=startClm, value=data)
                j -= 1
            else:
                ws.cell(row=startRow + i, column=startClm + j).font = font_wh
                ws.cell(row=startRow + i, column=startClm + j).alignment = al_cc
                ws.cell(row=startRow + i, column=startClm + j).fill = fill_bl
                ws.cell(row=startRow + i, column=startClm + j).border = border
                ws.cell(row=startRow + i, column=startClm + j, value=data)
            if tmp == data and startCell is None:
                startCell = ws.cell(row=startRow + i, column=startClm + j - 1).coordinate
            elif tmp != data and startCell is not None:
                # 横軸の結合
                ws.merge_cells(startCell + ':' + ws.cell(row=startRow + i, column=startClm + j - 1).coordinate)
                startCell = None
            tmp = data
    
    # 縦軸の結合
    for i in range(len(excel_header_list[0])):
        # 結合開始位置
        startCell = None
        # 前回ループ時の値との比較用
        tmp = None
        for j in range(depth):
            data = excel_header_list[j][i]
            if tmp == data:
                if startCell is None:
                    startCell = ws.cell(row=startRow + j - 1, column=startClm + i).coordinate
                if j == depth - 1:
                    # 縦軸の結合
                    ws.merge_cells(startCell + ':' + ws.cell(row=startRow + j, column=startClm + i).coordinate)
            elif tmp != data and startCell is not None:
                # 縦軸の結合
                ws.merge_cells(startCell + ':' + ws.cell(row=startRow + j, column=startClm + i).coordinate)
                startCell = None
            tmp = data
    return ws


# エクセルヘッダー部のカラム情報を作成する
def create_column_info(lang, ws, startRow, startClm, retList_t_common_menu_column_link, name_define_list, menu_table_link_list):
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # 背景色
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_ltw = Alignment(horizontal='left', vertical='top', wrapText=True)
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    sideDash = Side(border_style="dashed", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    borderDash = Border(left=side, right=side, top=sideDash, bottom=sideDash)
    
    # カラム位置調整用フラグ
    column_flg = False
    dataVaridationDict = {}
    column_class_file = ['9', '20']
    skip_cnt = 0
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        column_name = dict_menu_column.get('COLUMN_NAME_' + lang.upper())
        column_name_rest = dict_menu_column.get('COLUMN_NAME_REST')
        column_class = dict_menu_column.get('COLUMN_CLASS')
        auto_input = dict_menu_column.get('AUTO_INPUT')
        input_item = dict_menu_column.get('INPUT_ITEM')
        required_item = dict_menu_column.get('REQUIRED_ITEM')
        view_item = dict_menu_column.get('VIEW_ITEM')
        column_num = 0
        
        if input_item == '2' and view_item == '0':
            # excelに表示しない
            skip_cnt += 1
            continue

        if column_class == '23':
            # excelに表示しない
            skip_cnt += 1
            continue

        # 廃止フラグ
        msg = g.appmsg.get_api_message('MSG-30015')
        if column_name == msg:
                column_num = 4
                column_flg = True
        else:
            column_num = startClm + i + 1 - skip_cnt
            if column_flg:
                column_num -= 1
        
        # 項目名
        ws.cell(row=startRow + 6, column=column_num).font = font_wh
        ws.cell(row=startRow + 6, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 6, column=column_num).fill = fill_bl
        ws.cell(row=startRow + 6, column=column_num).border = border
        ws.cell(row=startRow + 6, column=column_num, value=column_name)
        
        # 項目名に合わせて幅の調整をする
        column_len = len(str(column_name))
        adjusted_width = (int(column_len) + 2) * 1.2
        ws.column_dimensions[get_column_letter(column_num)].width = adjusted_width
        
        # 最終更新日時はマイクロ秒部分を見えないように幅を設定する
        msg = g.appmsg.get_api_message('MSG-30017')
        if column_name == msg:
            ws.column_dimensions[get_column_letter(column_num)].width = 15.7
        
        # 明細1行目に入力規則を設定する
        if column_name_rest in name_define_list:
            dv = DataValidation(type='list', formula1=column_name_rest)
            dv.add(ws.cell(row=startRow + 7, column=column_num))
            ws.add_data_validation(dv)
            if column_name_rest not in dataVaridationDict:
                dataVaridationDict[get_column_letter(column_num)] = column_name_rest
        
        # 登録
        # 更新
        tmp = '×'
        if auto_input != '1' and input_item == '1' and required_item == '1' and column_class not in column_class_file:
            tmp = '●'
        elif auto_input != '1' and input_item == '1' and required_item == '0' and column_class not in column_class_file:
            tmp = '○'
        ws.cell(row=startRow, column=column_num).font = font_bl
        ws.cell(row=startRow, column=column_num).fill = fill_gr
        ws.cell(row=startRow, column=column_num).alignment = al_cc
        ws.cell(row=startRow, column=column_num).border = borderDash
        ws.cell(row=startRow, column=column_num, value=tmp)
        
        ws.cell(row=startRow + 1, column=column_num).font = font_bl
        ws.cell(row=startRow + 1, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 1, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 1, column=column_num).border = borderDash
        ws.cell(row=startRow + 1, column=column_num, value=tmp)
        
        tmp = '×'
        # カラムクラスが12（NoteColumn）の場合は任意を設定する
        if column_class == '12':
            tmp = '○'
        # 廃止(×固定)
        ws.cell(row=startRow + 2, column=column_num).font = font_bl
        ws.cell(row=startRow + 2, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 2, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 2, column=column_num).border = borderDash
        ws.cell(row=startRow + 2, column=column_num, value=tmp)
        
        # 復活(×固定)
        ws.cell(row=startRow + 3, column=column_num).font = font_bl
        ws.cell(row=startRow + 3, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 3, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 3, column=column_num).border = borderDash
        ws.cell(row=startRow + 3, column=column_num, value=tmp)
        
        # 改行
        # 改行OK：MultiTextColumn、NoteColumn、SensitiveMultiTextColumnのもの
        # 改行NG：上記以外
        new_line = 'NG'
        if column_class == '2' or column_class == '12' or column_class == '17':
            new_line = 'OK'
        ws.cell(row=startRow + 4, column=column_num).font = font_bl
        ws.cell(row=startRow + 4, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 4, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 4, column=column_num).border = border
        ws.cell(row=startRow + 4, column=column_num, value=new_line)
        
        # その他注意事項
        ws.cell(row=startRow + 5, column=column_num).font = font_bl
        ws.cell(row=startRow + 5, column=column_num).fill = fill_gr
        ws.cell(row=startRow + 5, column=column_num).alignment = al_ltw
        ws.cell(row=startRow + 5, column=column_num).border = border
        ws.cell(row=startRow + 5, column=column_num, value=dict_menu_column.get('DESCRIPTION_' + lang.upper()))
    
    # フィルター設定
    st = ws.cell(row=startRow + 6, column=3).coordinate
    ed = ws.cell(row=startRow + 6, column=ws.max_column).coordinate
    ws.auto_filter.ref = st + ':' + ed
    
    # セルの高さの設定
    ws.row_dimensions[startRow + 4].height = 13.5
    ws.row_dimensions[startRow + 5].height = 85.5
    
    return ws, dataVaridationDict


# エクセルヘッダー部のカラム情報を作成する(変更履歴ver)
def create_column_info_trace_history(lang, ws, startRow, startClm, retList_t_common_menu_column_link):
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # 背景色
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_ltw = Alignment(horizontal='left', vertical='top', wrapText=True)
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # カラム位置調整用フラグ
    column_flg = False
    # 登録不可の行を記憶しておく
    gray_column = []
    skip_cnt = 0
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        column_name = dict_menu_column.get('COLUMN_NAME_' + lang.upper())
        auto_input = dict_menu_column.get('AUTO_INPUT')
        input_item = dict_menu_column.get('INPUT_ITEM')
        view_item = dict_menu_column.get('VIEW_ITEM')
        column_class = dict_menu_column.get('COLUMN_CLASS')
        column_num = 0
        
        if input_item == '2' and view_item == '0':
            # excelに表示しない
            skip_cnt += 1
            continue

        if column_class == '23':
            # excelに表示しない
            skip_cnt += 1
            continue

        # 廃止フラグ
        msg = g.appmsg.get_api_message('MSG-30015')
        if column_name == msg:
            column_num = 3
            column_flg = True
        else:
            column_num = startClm + i + 1 - skip_cnt
            if column_flg:
                column_num -= 1
        
        # 項目名
        ws.cell(row=startRow + 1, column=column_num).font = font_wh
        ws.cell(row=startRow + 1, column=column_num).alignment = al_cc
        ws.cell(row=startRow + 1, column=column_num).fill = fill_bl
        ws.cell(row=startRow + 1, column=column_num).border = border
        ws.cell(row=startRow + 1, column=column_num, value=column_name)
        
        # 項目名に合わせて幅の調整をする
        column_len = len(str(column_name))
        adjusted_width = (int(column_len) + 2) * 1.2
        ws.column_dimensions[get_column_letter(column_num)].width = adjusted_width
        
        # 最終更新日時はマイクロ秒部分を見えないように幅を設定する
        msg = g.appmsg.get_api_message('MSG-30017')
        if column_name == msg:
            ws.column_dimensions[get_column_letter(column_num)].width = 15.7
        
        # その他注意事項
        ws.cell(row=startRow, column=column_num).font = font_bl
        ws.cell(row=startRow, column=column_num).fill = fill_gr
        ws.cell(row=startRow, column=column_num).alignment = al_ltw
        ws.cell(row=startRow, column=column_num).border = border
        ws.cell(row=startRow, column=column_num, value=dict_menu_column.get('DESCRIPTION_' + lang.upper()))
        
        # 最後に列をグレーにするために登録不可の行を記憶しておく
        if auto_input == '1' or input_item == '0':
            gray_column.append(get_column_letter(column_num))
    
    # フィルター設定
    st = ws.cell(row=startRow + 1, column=3).coordinate
    ed = ws.cell(row=startRow + 1, column=ws.max_column).coordinate
    ws.auto_filter.ref = st + ':' + ed
    
    # セルの高さの設定
    ws.row_dimensions[startRow].height = 13.5
    ws.row_dimensions[startRow].height = 85.5
    
    return ws, gray_column


# フッターを作成する
def create_footer(ws, font_wh, al_lt, al_cc, fill_bl):
    # 罫線
    side = Side(border_style="thin", color="000000")
    border_r = Border(left=side, right=None, top=side, bottom=side)
    border_l = Border(left=None, right=side, top=side, bottom=side)
    border_rl = Border(left=None, right=None, top=side, bottom=side)
    
    # フッター
    footer_row = ws.max_row + 1
    for i in range(ws.max_column):
        if i == 0:
            msg = g.appmsg.get_api_message('MSG-30011')
            ws.cell(row=footer_row, column=1).font = font_wh
            ws.cell(row=footer_row, column=1).alignment = al_lt
            ws.cell(row=footer_row, column=1).fill = fill_bl
            ws.cell(row=footer_row, column=1).border = border_r
            ws.cell(row=footer_row, column=1, value=msg)
        elif i == ws.max_column - 1:
            ws.cell(row=footer_row, column=i + 1).font = font_wh
            ws.cell(row=footer_row, column=i + 1).alignment = al_cc
            ws.cell(row=footer_row, column=i + 1).fill = fill_bl
            ws.cell(row=footer_row, column=i + 1).border = border_l
            ws.cell(row=footer_row, column=i + 1, value='')
        else:
            ws.cell(row=footer_row, column=i + 1).font = font_wh
            ws.cell(row=footer_row, column=i + 1).alignment = al_cc
            ws.cell(row=footer_row, column=i + 1).fill = fill_bl
            ws.cell(row=footer_row, column=i + 1).border = border_rl
            ws.cell(row=footer_row, column=i + 1, value='')
    
    return ws


# 明細一行目の書式設定
def detail_first_line_format(ws, startRow, dataVaridationDict):
    # 背景色
    fill_wh = PatternFill(fill_type='solid', fgColor='FFFFFF')
    
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    ws.cell(row=startRow + 7, column=1).fill = fill_wh
    ws.cell(row=startRow + 7, column=1, value='')
    
    ws.cell(row=startRow + 7, column=2).fill = fill_wh
    ws.cell(row=startRow + 7, column=2, value='')
    
    # 実行処理種別
    ws.cell(row=startRow + 7, column=3).font = font_bl
    ws.cell(row=startRow + 7, column=3).alignment = al_cc
    ws.cell(row=startRow + 7, column=3).border = border
    ws.cell(row=startRow + 7, column=3, value='-')
    
    # 入力規則の設定
    dv = DataValidation(type='list', formula1='FILTER_ROW_EDIT_BY_FILE')
    dv.add(ws.cell(row=startRow + 7, column=3))
    ws.add_data_validation(dv)
    dataVaridationDict[get_column_letter(3)] = 'FILTER_ROW_EDIT_BY_FILE'
    
    column_num = ws.max_column + 1
    for col_i in range(4, column_num):
        ws.cell(row=startRow + 7, column=col_i).number_format = openpyxl.styles.numbers.FORMAT_TEXT
        ws.cell(row=startRow + 7, column=col_i).font = font_bl
        ws.cell(row=startRow + 7, column=col_i).border = border
        ws.cell(row=startRow + 7, column=col_i, value='')
    
    return ws, dataVaridationDict


# 明細一行目の書式設定(変更履歴ver)
def detail_first_line_format_trace_history(ws, startRow, dataVaridationDict):
    # 背景色
    fill_wh = PatternFill(fill_type='solid', fgColor='FFFFFF')
    
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # 入力規則の設定
    dv = DataValidation(type='list', formula1='FILTER_ROW_EDIT_BY_FILE')
    dv.add(ws.cell(row=startRow, column=3))
    ws.add_data_validation(dv)
    dataVaridationDict[get_column_letter(3)] = 'FILTER_ROW_EDIT_BY_FILE'
    
    column_num = ws.max_column + 1
    for col_i in range(1, column_num):
        # ws.cell(row=startRow, column=col_i).number_format = openpyxl.styles.numbers.FORMAT_TEXT
        ws.cell(row=startRow, column=col_i).fill = fill_wh
        ws.cell(row=startRow, column=col_i).font = font_bl
        ws.cell(row=startRow, column=col_i).border = border
        ws.cell(row=startRow, column=col_i, value='')
    
    return ws, dataVaridationDict


# 空行を作成する
def create_blank_line(ws, dataVaridationDict, addline, trace_history_flag=False):
    # 空行追加処理
    min_col = 1
    min_row = ws.max_row
    max_col = ws.max_column
    max_row = ws.max_row + addline
    shift_col = 0
    shift_row = 1
    for col in range(min_col, max_col + 1):
        for row in range(min_row, max_row + 1):
            # コピー元のコードを作成(column = 1, row = 1 → A1)
            copyFrmCoord = get_column_letter(col) + str(row)
            
            # コピー先のコードを作成
            copyToCol = get_column_letter(col + shift_col)
            copyToCoord = copyToCol + str(row + shift_row)
            
            # コピー先に値をコピー
            if not trace_history_flag and copyToCol == 'C':
                ws[copyToCoord].value = ws[copyFrmCoord].value
            
            if copyToCol in dataVaridationDict:
                dv = DataValidation(type='list', formula1=dataVaridationDict[copyToCol])
                dv.add(copyToCoord)
                ws.add_data_validation(dv)
            
            # 書式がある場合は書式もコピー
            if ws[copyFrmCoord].has_style:
                ws[copyToCoord]._style = copy(ws[copyFrmCoord]._style)

    return ws


def collect_excel_format(objdbca, organization_id, workspace_id, menu, menu_record, menu_table_link_record):
    """
        新規登録用Excelを取得する

        ARGS:
            objdbca: DB接クラス  DBConnectWs()
            organization_id: Organization ID
            workspace_id: Workspace ID
            menu: メニュー名
            menu_record: メニュー管理のレコード
            menu_table_link_record: メニュー-テーブル紐付管理のレコード
        RETURN:
            Excelをbase64した値
    """
    # 言語設定取得
    # 変数定義
    lang = g.LANGUAGE
    
    # make storage directory for excel
    strage_path = os.environ.get('STORAGEPATH')
    excel_dir = strage_path + "/".join([organization_id, workspace_id]) + "/tmp/excel"
    if not os.path.isdir(excel_dir):
        os.makedirs(excel_dir)
        g.applogger.debug("made excel_dir")
    
    # テーブル名
    t_common_column_group = 'T_COMN_COLUMN_GROUP'
    t_common_menu_column_link = 'T_COMN_MENU_COLUMN_LINK'
    
    # メニューのカラム情報を取得
    objmenu = load_table.loadTable(objdbca, menu)   # noqa: F405
    mode = 'excel'
    result_data = objmenu.rest_filter({}, mode)
    
    #### result_code,msg未対応
    result = {
        "result": "result_code", #result_data[0],
        "data": result_data, #result_data[1],
        "message": "msg" #result_data[2]
    }
    
    # 対象メニューを特定するためのIDを取得する
    menu_id = menu_record[0].get('MENU_ID')
    menu_name = menu_record[0].get('MENU_NAME_' + lang.upper())
    file_name = menu_name + '_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
    file_path = excel_dir + '/' + file_name
    
    # 『メニュー-カラム紐付管理』テーブルから対象のデータを取得
    retList_t_common_menu_column_link = objdbca.table_select(t_common_menu_column_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s ORDER BY COLUMN_DISP_SEQ ASC', [menu_id, 0])
    if not retList_t_common_menu_column_link:
        log_msg_args = [menu]
        api_msg_args = [menu]
        raise AppException("499-00005", log_msg_args, api_msg_args)  # noqa: F405
    
    # 使用するCOL_GROUP_IDだけを配列に確保しておく
    active_col_group_id = ['']
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        tmp = dict_menu_column.get('COL_GROUP_ID')
        if tmp is not None and tmp not in active_col_group_id:
            active_col_group_id.append(tmp)
    
    # IDColumn項目のプルダウン一覧の取得
    pulldown_list = menu_info.collect_pulldown_list(objdbca, menu, menu_record)
    
    # 登録更新廃止復活フラグを取得する
    menu_table_link_list = []
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_INSERT_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_UPDATE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_DISUSE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_REUSE_FLAG'))
    
    # 色の設定
    # 背景色
    fill_wh = PatternFill(fill_type='solid', fgColor='FFFFFF')
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_lt = Alignment(horizontal='left', vertical='top')
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    sideDash = Side(border_style="dashed", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    borderDash = Border(left=side, right=side, top=sideDash, bottom=sideDash)
    
    # ワークブックの新規作成と保存
    wb = Workbook()
    
    # マスタ
    msg = g.appmsg.get_api_message('MSG-30012')
    wb.create_sheet(msg)
    # フィルタ条件
    msg = g.appmsg.get_api_message('MSG-30013')
    wb.create_sheet(msg)
    
    # 「マスタ」シート作成
    # 「マスタ」シートの名前の定義をメインのシートの入力規則で使用するため「マスタ」シートから作成する
    name_define_list = make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, menu_table_link_list, pulldown_list)
    
    # シートを指定して編集
    ws = wb.active
    ws.title = menu_name[:31]
    
    # ヘッダー部編集
    startRow = 1
    startClm = 4
    
    # 『カラムグループ管理』テーブルからカラムグループ一覧を取得
    ret = objdbca.table_select(t_common_column_group, 'WHERE COL_GROUP_ID IN %s AND DISUSE_FLAG = %s', [active_col_group_id, 0])
    dict_column_group_id = {}
    dict_column_group_id_name = {}
    if ret:
        for recode in ret:
            dict_column_group_id[recode.get('COL_GROUP_ID')] = recode.get('PA_COL_GROUP_ID')
            # IDと名前の辞書も作成する
            dict_column_group_id_name[recode.get('COL_GROUP_ID')] = recode.get('COL_GROUP_NAME_' + lang.upper())
    
    # ヘッダーの階層の深さを調べる
    depth = get_col_group_depth(retList_t_common_menu_column_link, dict_column_group_id)
    
    # エクセルに表示するヘッダー項目を二次元配列に構築する
    excel_header_list, header_order = create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id, menu_table_link_list)
    
    # 1行目（項目名）のヘッダーを作成する
    ws = create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border)
    
    startRow = startRow + depth
    startDetailRow = startRow + 7
    
    # 2行目以降を作成する
    # 固定部分
    ws = make_template(ws, startRow, font_bl, fill_gr, borderDash, depth)
    
    # エクセルヘッダー部のカラム情報を作成する
    ws, dataVaridationDict = create_column_info(lang, ws, startRow, startClm, retList_t_common_menu_column_link, name_define_list, menu_table_link_list)
    
    # 明細部編集
    # 明細1行目編集
    ws, dataVaridationDict = detail_first_line_format(ws, startRow, dataVaridationDict)
    
    # ウィンドウ枠の固定
    ws.freeze_panes = ws.cell(row=startRow + 7, column=4).coordinate
    
    # 空行追加処理
    ws = create_blank_line(ws, dataVaridationDict, 9)
    
    # 登録が×の列をグレーにする
    # 明細の数を求める
    column_num = ws.max_column + 1
    row_num = ws.max_row - startDetailRow + 1
    for col_i in range(4, column_num):
        for row_j in range(row_num):
            if ws.cell(row=depth + 1, column=col_i).value == '×':
                # セルの背景色をグレーにする
                ws.cell(row=startDetailRow + row_j, column=col_i).fill = fill_gr
    
    # フッターを作成する
    ws = create_footer(ws, font_wh, al_lt, al_cc, fill_bl)
    
    # フィルタ条件シートを指定して編集
    # フィルタ条件
    msg = g.appmsg.get_api_message('MSG-30013')
    ws_filter = wb[msg]
    # 出力日時
    msg = g.appmsg.get_api_message('MSG-30014')
    ws_filter['A1'] = msg
    # 廃止
    msg = g.appmsg.get_api_message('MSG-30006')
    ws_filter['D1'] = msg
    ws_filter['A2'] = datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
    # 全レコード
    msg = g.appmsg.get_api_message('MSG-30018')
    ws_filter['D2'] = msg
    
    wb.save(file_path)  # noqa: E303
    
    # 編集してきたエクセルファイルをエンコードする
    wbEncode = file_encode(file_path)  # noqa: F405 F841
    # エンコード後wbは削除する
    os.remove(file_path)
    
    return wbEncode


def collect_excel_journal(objdbca, organization_id, workspace_id, menu, menu_record, menu_table_link_record):
    """
        変更履歴のExcelを取得する

        ARGS:
            objdbca: DB接クラス  DBConnectWs()
            organization_id: Organization ID
            workspace_id: Workspace ID
            menu: メニュー名
            menu_record: メニュー管理のレコード
            menu_table_link_record: メニュー-テーブル紐付管理のレコード
        RETURN:
            Excelをbase64した値
    """
    # 言語設定取得
    lang = g.LANGUAGE
    
    # make storage directory for excel
    strage_path = os.environ.get('STORAGEPATH')
    excel_dir = strage_path + "/".join([organization_id, workspace_id]) + "/tmp/excel"
    if not os.path.isdir(excel_dir):
        os.makedirs(excel_dir)
        g.applogger.debug("made excel_dir")
    
    # テーブル名
    t_common_column_group = 'T_COMN_COLUMN_GROUP'
    t_common_menu_column_link = 'T_COMN_MENU_COLUMN_LINK'
    
    mode = 'excel_jnl_all'
    objmenu = load_table.loadTable(objdbca, menu)   # noqa: F405
    if objmenu.get_objtable() is False:
        log_msg_args = ["not menu or table"]
        api_msg_args = ["not menu or table"]
        raise AppException("401-00001", log_msg_args, api_msg_args)     # noqa: F405

    status_code, result, msg = objmenu.rest_filter({}, mode)
    if status_code != '000-00000':
        log_msg_args = [msg]
        api_msg_args = [msg]
        raise AppException(status_code, log_msg_args, api_msg_args)     # noqa: F405
    
    # 変更履歴
    msg = g.appmsg.get_api_message('MSG-30022')
    # 対象メニューを特定するためのIDを取得する
    menu_id = menu_record[0].get('MENU_ID')
    menu_name = menu_record[0].get('MENU_NAME_' + lang.upper())
    file_name = menu_name + '_' + msg + '_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
    file_path = excel_dir + '/' + file_name
    
    # 『メニュー-カラム紐付管理』テーブルから対象のデータを取得
    retList_t_common_menu_column_link = objdbca.table_select(t_common_menu_column_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s ORDER BY COLUMN_DISP_SEQ ASC', [menu_id, 0])
    if not retList_t_common_menu_column_link:
        log_msg_args = [menu]
        api_msg_args = [menu]
        raise AppException("499-00005", log_msg_args, api_msg_args)  # noqa: F405
    
    # 使用するCOL_GROUP_IDだけを配列に確保しておく
    active_col_group_id = ['']
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        tmp = dict_menu_column.get('COL_GROUP_ID')
        if tmp is not None and tmp not in active_col_group_id:
            active_col_group_id.append(tmp)
    
    # IDColumn項目のプルダウン一覧の取得
    pulldown_list = menu_info.collect_pulldown_list(objdbca, menu, menu_record)
    
    # 登録更新廃止復活フラグを取得する
    menu_table_link_list = []
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_INSERT_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_UPDATE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_DISUSE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_REUSE_FLAG'))
    
    # 色の設定
    # 背景色
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    
    # ワークブックの新規作成と保存
    wb = Workbook()
    
    # マスタ
    msg = g.appmsg.get_api_message('MSG-30012')
    wb.create_sheet(msg)
    
    # 「マスタ」シート作成
    # 「マスタ」シートの名前の定義をメインのシートの入力規則で使用するためまず「マスタ」シートから作成する
    name_define_list = make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, menu_table_link_list, pulldown_list)
    
    # シートを指定して編集
    ws = wb.active
    ws.title = menu_name[:31]
    
    # ヘッダー部編集
    startRow = 1
    startClm = 3
    
    # 『カラムグループ管理』テーブルからカラムグループ一覧を取得
    ret = objdbca.table_select(t_common_column_group, 'WHERE COL_GROUP_ID IN %s AND DISUSE_FLAG = %s', [active_col_group_id, 0])
    dict_column_group_id = {}
    dict_column_group_id_name = {}
    if ret:
        for recode in ret:
            dict_column_group_id[recode.get('COL_GROUP_ID')] = recode.get('PA_COL_GROUP_ID')
            # IDと名前の辞書も作成する
            dict_column_group_id_name[recode.get('COL_GROUP_ID')] = recode.get('COL_GROUP_NAME_' + lang.upper())
    
    # ヘッダーの階層の深さを調べる
    depth = get_col_group_depth(retList_t_common_menu_column_link, dict_column_group_id)
    
    # エクセルに表示するヘッダー項目を二次元配列に構築する
    excel_header_list, header_order = create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id, menu_table_link_list)
    
    # 1行目（項目名）のヘッダーを作成する
    ws = create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border)
    
    startRow = startRow + depth
    startDetailRow = startRow + 2
    
    # 2行目以降を作成する
    # 固定部分
    ws = make_template_trace_history(ws, depth)
    
    # エクセルヘッダー部のカラム情報を作成する
    ws, gray_column = create_column_info_trace_history(lang, ws, startRow, startClm, retList_t_common_menu_column_link)
    
    # 明細部編集
    # 明細1行目編集
    dataVaridationDict = {}
    ws, dataVaridationDict = detail_first_line_format_trace_history(ws, startDetailRow, dataVaridationDict)
    
    # ウィンドウ枠の固定
    ws.freeze_panes = ws.cell(row=startDetailRow, column=1).coordinate
    
    for param in result:
        for k, v in param.items():
            if k == 'parameter':
                for i, (key, value) in enumerate(v.items()):
                    if i == 0:
                        # 廃止フラグ
                        ws.cell(row=startDetailRow, column=3).font = font_bl
                        ws.cell(row=startDetailRow, column=3).alignment = al_cc
                        ws.cell(row=startDetailRow, column=3).border = border
                        ws.cell(row=startDetailRow, column=3, value='')
                        # 入力規則の設定
                        dv = DataValidation(type='list', formula1='FILTER_ROW_EDIT_BY_FILE')
                        dv.add(ws.cell(row=startDetailRow, column=3))
                        ws.add_data_validation(dv)
                        dataVaridationDict[get_column_letter(3)] = 'FILTER_ROW_EDIT_BY_FILE'
                    
                    if key == 'journal_action':
                        continue
                    elif key == 'discard':
                        if value == '1':
                            # 廃止
                            msg = g.appmsg.get_api_message('MSG-30006')
                            ws.cell(row=startDetailRow, column=3, value=msg)
                    else:
                        if i < 3:
                            column_num = i + 1
                        else:
                            if key in header_order:
                                column_num = header_order.index(key) + 3
                            else:
                                continue

                        # dict or list 形式を json.dumps
                        if isinstance(value, dict) or isinstance(value, list):
                            value = json.dumps(value, ensure_ascii=False)
                            
                        ws.cell(row=startDetailRow, column=column_num).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                        ws.cell(row=startDetailRow, column=column_num).font = font_bl
                        ws.cell(row=startDetailRow, column=column_num).border = border
                        ws.cell(row=startDetailRow, column=column_num, value=value)
                        if key == 'journal_datetime':
                            ws.cell(row=startDetailRow, column=i + 1).alignment = Alignment(horizontal='fill')
                        
                        if key in name_define_list:
                            dv = DataValidation(type='list', formula1=key)
                            dv.add(ws.cell(row=startDetailRow, column=column_num))
                            ws.add_data_validation(dv)
                            if key not in dataVaridationDict:
                                dataVaridationDict[get_column_letter(column_num)] = key
                startDetailRow += 1
    
    # 空行追加処理
    ws = create_blank_line(ws, dataVaridationDict, 10, True)
    
    # 登録が×の列をグレーにする
    # 明細の数を求める
    column_num = ws.max_column + 1
    row_num = ws.max_row - (startRow + 1)
    for col_i in range(3, column_num):
        for row_j in range(row_num):
            if get_column_letter(col_i) in gray_column:
                # セルの背景色をグレーにする
                ws.cell(row=startRow + row_j + 2, column=col_i).fill = fill_gr
    
    wb.save(file_path)  # noqa: E303
    
    # 編集してきたエクセルファイルをエンコードする
    wbEncode = file_encode(file_path)  # noqa: F405 F841
    # エンコード後wbは削除する
    os.remove(file_path)    # noqa: F405
    
    return wbEncode


def collect_excel_filter(objdbca, organization_id, workspace_id, menu, menu_record, menu_table_link_record, filter_parameter):
    """
        検索条件を指定し、Excelを取得する

        ARGS:
            objdbca: DB接クラス  DBConnectWs()
            organization_id: Organization ID
            workspace_id: Workspace ID
            menu: メニュー名
            menu_record: メニュー管理のレコード
            menu_table_link_record: メニュー-テーブル紐付管理のレコード
            filter_parameter: 検索条件
        RETURN:
            Excelをbase64した値
    """
    # 言語設定取得
    lang = g.LANGUAGE
    
    # make storage directory for excel
    strage_path = os.environ.get('STORAGEPATH')
    excel_dir = strage_path + "/".join([organization_id, workspace_id]) + "/tmp/excel"
    if not os.path.isdir(excel_dir):
        os.makedirs(excel_dir)
        g.applogger.debug("made excel_dir")
    
    # テーブル名
    t_common_column_group = 'T_COMN_COLUMN_GROUP'
    t_common_menu_column_link = 'T_COMN_MENU_COLUMN_LINK'
    
    mode = 'excel'
    objmenu = load_table.loadTable(objdbca, menu)   # noqa: F405
    if objmenu.get_objtable() is False:
        log_msg_args = ["not menu or table"]
        api_msg_args = ["not menu or table"]
        raise AppException("401-00001", log_msg_args, api_msg_args) # noqa: F405
    
    status_code, result, msg = objmenu.rest_filter(filter_parameter, mode)
    if status_code != '000-00000':
        log_msg_args = [msg]
        api_msg_args = [msg]
        raise AppException(status_code, log_msg_args, api_msg_args) # noqa: F405
    
    # 対象メニューを特定するためのIDを取得する
    menu_id = menu_record[0].get('MENU_ID')
    menu_name = menu_record[0].get('MENU_NAME_' + lang.upper())
    file_name = menu_name + '_' + datetime.datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
    file_path = excel_dir + '/' + file_name
    
    # 『メニュー-カラム紐付管理』テーブルから対象のデータを取得
    retList_t_common_menu_column_link = objdbca.table_select(t_common_menu_column_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s ORDER BY COLUMN_DISP_SEQ ASC', [menu_id, 0])
    if not retList_t_common_menu_column_link:
        log_msg_args = [menu]
        api_msg_args = [menu]
        raise AppException("499-00005", log_msg_args, api_msg_args)  # noqa: F405

    # 使用するCOL_GROUP_IDだけを配列に確保しておく
    active_col_group_id = ['']
    for i, dict_menu_column in enumerate(retList_t_common_menu_column_link):
        tmp = dict_menu_column.get('COL_GROUP_ID')
        if tmp is not None and tmp not in active_col_group_id:
            active_col_group_id.append(tmp)
    
    # IDColumn項目のプルダウン一覧の取得
    pulldown_list = menu_info.collect_pulldown_list(objdbca, menu, menu_record)
    
    # 登録更新廃止復活フラグを取得する
    menu_table_link_list = []
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_INSERT_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_UPDATE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_DISUSE_FLAG'))
    menu_table_link_list.append(menu_table_link_record[0].get('ROW_REUSE_FLAG'))
    
    # 色の設定
    # 背景色
    fill_wh = PatternFill(fill_type='solid', fgColor='FFFFFF')
    fill_bl = PatternFill(fill_type='solid', fgColor='00459D')
    fill_gr = PatternFill(fill_type='solid', fgColor='D9D9D9')
    
    # 文字色
    font_bl = openpyxl.styles.Font(name='メイリオ', size=8, color='000000')
    font_wh = openpyxl.styles.Font(name='メイリオ', size=8, color='FFFFFF')
    
    # Alignmentの設定
    al_cc = Alignment(horizontal='center', vertical='center')
    al_lt = Alignment(horizontal='left', vertical='top')
    
    # 罫線の設定
    side = Side(border_style="thin", color="000000")
    sideDash = Side(border_style="dashed", color="000000")
    border = Border(left=side, right=side, top=side, bottom=side)
    borderDash = Border(left=side, right=side, top=sideDash, bottom=sideDash)
    
    # ワークブックの新規作成と保存
    wb = Workbook()
    
    # マスタ
    msg = g.appmsg.get_api_message('MSG-30012')
    wb.create_sheet(msg)
    # フィルタ条件
    msg = g.appmsg.get_api_message('MSG-30013')
    wb.create_sheet(msg)
    
    # 「マスタ」シート作成
    # 「マスタ」シートの名前の定義をメインのシートの入力規則で使用するため「マスタ」シートから作成する
    name_define_list = make_master_sheet(wb, objdbca, lang, retList_t_common_menu_column_link, menu_table_link_list, pulldown_list)
    
    # シートを指定して編集
    ws = wb.active
    ws.title = menu_name[:31]
    
    # ヘッダー部編集
    startRow = 1
    startClm = 4
    
    # 『カラムグループ管理』テーブルからカラムグループ一覧を取得
    ret = objdbca.table_select(t_common_column_group, 'WHERE COL_GROUP_ID IN %s AND DISUSE_FLAG = %s', [active_col_group_id, 0])
    dict_column_group_id = {}
    dict_column_group_id_name = {}
    if ret:
        for recode in ret:
            dict_column_group_id[recode.get('COL_GROUP_ID')] = recode.get('PA_COL_GROUP_ID')
            # IDと名前の辞書も作成する
            dict_column_group_id_name[recode.get('COL_GROUP_ID')] = recode.get('COL_GROUP_NAME_' + lang.upper())
    
    # ヘッダーの階層の深さを調べる
    depth = get_col_group_depth(retList_t_common_menu_column_link, dict_column_group_id)
    
    # エクセルに表示するヘッダー項目を二次元配列に構築する
    excel_header_list, header_order = create_excel_headerlist(lang, ws, depth, retList_t_common_menu_column_link, dict_column_group_id_name, dict_column_group_id, menu_table_link_list)
    
    # 1行目（項目名）のヘッダーを作成する
    ws = create_excel_header_firstline(ws, excel_header_list, depth, startRow, startClm, font_wh, al_cc, fill_bl, border)
    
    startRow = startRow + depth
    startDetailRow = startRow + 7
    
    # 2行目以降を作成する
    # 固定部分
    ws = make_template(ws, startRow, font_bl, fill_gr, borderDash, depth)
    
    # エクセルヘッダー部のカラム情報を作成する
    ws, dataVaridationDict = create_column_info(lang, ws, startRow, startClm, retList_t_common_menu_column_link, name_define_list, menu_table_link_list)
    
    # 明細部編集
    # 明細1行目編集
    ws, dataVaridationDict = detail_first_line_format(ws, startRow, dataVaridationDict)
    
    # ウィンドウ枠の固定
    ws.freeze_panes = ws.cell(row=startRow + 7, column=4).coordinate
    
    for param in result:
        for k, v in param.items():
            if k == 'parameter':
                # # カラム位置調整用フラグ
                # column_flg = False
                for i, (key, value) in enumerate(v.items()):
                    if i == 0:
                        ws.cell(row=startRow + 7, column=1).fill = fill_wh
                        ws.cell(row=startRow + 7, column=1, value='')
                        
                        ws.cell(row=startRow + 7, column=2).fill = fill_wh
                        ws.cell(row=startRow + 7, column=2, value='')
                        
                        # 実行処理種別
                        ws.cell(row=startRow + 7, column=3).font = font_bl
                        ws.cell(row=startRow + 7, column=3).alignment = al_cc
                        ws.cell(row=startRow + 7, column=3).border = border
                        ws.cell(row=startRow + 7, column=3, value='-')
                        # 入力規則の設定
                        dv = DataValidation(type='list', formula1='FILTER_ROW_EDIT_BY_FILE')
                        dv.add(ws.cell(row=startRow + 7, column=3))
                        ws.add_data_validation(dv)
                        dataVaridationDict[get_column_letter(3)] = 'FILTER_ROW_EDIT_BY_FILE'
                    
                    if key == 'discard':
                        # 廃止
                        msg = g.appmsg.get_api_message('MSG-30006')
                        if value == '1':
                            value = msg
                        else:
                            value = ''

                    # dict or list 形式を json.dumps
                    if isinstance(value, dict) or isinstance(value, list):
                        value = json.dumps(value, ensure_ascii=False)

                    if key in header_order:
                        column_num = header_order.index(key) + 4
                    else:
                        continue
                    
                    ws.cell(row=startRow + 7, column=column_num).number_format = openpyxl.styles.numbers.FORMAT_TEXT
                    ws.cell(row=startRow + 7, column=column_num).font = font_bl
                    ws.cell(row=startRow + 7, column=column_num).border = border
                    ws.cell(row=startRow + 7, column=column_num, value=value)
                    if key in name_define_list:
                        dv = DataValidation(type='list', formula1=key)
                        dv.add(ws.cell(row=startRow + 7, column=column_num))
                        ws.add_data_validation(dv)
                        if key not in dataVaridationDict:
                            dataVaridationDict[get_column_letter(column_num)] = key
                startRow = startRow + 1
    
    # 空行追加処理
    ws = create_blank_line(ws, dataVaridationDict, 10)
    
    # 登録が×の列をグレーにする
    # 明細の数を求める
    column_num = ws.max_column + 1
    row_num = ws.max_row - startDetailRow + 1
    for col_i in range(4, column_num):
        for row_j in range(row_num):
            if ws.cell(row=depth + 1, column=col_i).value == '×':
                # セルの背景色をグレーにする
                ws.cell(row=startDetailRow + row_j, column=col_i).fill = fill_gr
    
    # フッターを作成する
    ws = create_footer(ws, font_wh, al_lt, al_cc, fill_bl)
    
    # フィルタ条件シートを指定して編集
    # フィルタ条件
    msg = g.appmsg.get_api_message('MSG-30013')
    ws_filter = wb[msg]
    # 出力日時
    msg = g.appmsg.get_api_message('MSG-30014')
    ws_filter['A1'] = msg
    # 廃止
    msg = g.appmsg.get_api_message('MSG-30006')
    ws_filter['D1'] = msg
    ws_filter['A2'] = datetime.datetime.now().strftime('%Y/%m/%d %H:%M:%S')
    
    for i, (key, value) in enumerate(filter_parameter.items()):
        tmp = list(value.values())[0]
        filter_name = key
        if key in header_order:
            ind = header_order.index(key)
            filter_name = excel_header_list[-1][ind]
        ws_filter.cell(row=1, column=4 + i, value=filter_name)
        if key.lower() == 'discard':
            if tmp == '0':
                # 廃止含まず
                msg = g.appmsg.get_api_message('MSG-30016')
                ws_filter.cell(row=2, column=4 + i, value=msg)
            elif tmp == '1':
                # 廃止のみ
                msg = g.appmsg.get_api_message('MSG-30019')
                ws_filter.cell(row=2, column=4 + i, value=msg)
            else:
                # 全レコード
                msg = g.appmsg.get_api_message('MSG-30018')
                ws_filter.cell(row=2, column=4 + i, value=msg)
        else:
            if type(tmp) is list:
                for j, val in enumerate(tmp):
                    ws_filter.cell(row=2 + j, column=4 + i, value=tmp[j])
            elif type(tmp) is dict:
                filter_range = tmp.get('START', '') + '～' + tmp.get('END', '')
                ws_filter.cell(row=2, column=4 + i, value=filter_range)
    
    wb.save(file_path)  # noqa: E303
    
    # 編集してきたエクセルファイルをエンコードする
    wbEncode = file_encode(file_path)  # noqa: F405 F841
    # エンコード後wbは削除する
    os.remove(file_path)    # noqa: F405
    
    return wbEncode


def execute_excel_maintenance(objdbca, organization_id, workspace_id, menu, menu_record, excel_data):
    """execute_excel_maintenance

        Excelでレコードを登録/更新/廃止/復活する # noqa: E501

        ARGS:
            objdbca: DB接クラス  DBConnectWs()
            organization_id: Organization ID
            workspace_id: Workspace ID
            menu: メニュー名
            menu_record: メニュー管理のレコード
            excel_data: Excelをbase64した値
        RETURN:
            処理件数
    """
    # 言語設定取得
    lang = g.LANGUAGE

    # make storage directory for excel
    strage_path = os.environ.get('STORAGEPATH')
    excel_dir = strage_path + "/".join([organization_id, workspace_id]) + "/tmp/excel"
    if not os.path.isdir(excel_dir):
        os.makedirs(excel_dir)
        g.applogger.debug("made excel_dir")
    
    # 変数定義
    t_common_menu_column_link = 'T_COMN_MENU_COLUMN_LINK'
    
    # メニューのカラム情報を取得
    objmenu = load_table.loadTable(objdbca, menu)   # noqa: F405
    mode = 'excel'
    status_code, result, msg = objmenu.rest_filter({}, mode)
    if status_code != '000-00000':
        log_msg_args = [msg]
        api_msg_args = [msg]
        raise AppException(status_code, log_msg_args, api_msg_args)     # noqa: F405
    
    wbDecode = base64.b64decode(excel_data.encode('utf-8'))
    
    # 受け取ったデータを編集用として一時的にエクセルファイルに保存
    file_name = 'post_excel_maintenance_tmp.xlsx'
    file_path = excel_dir + '/' + file_name
    with open(file_path, "wb") as f:
        f.write(wbDecode)
    
    try:
        # ファイルを読み込む
        wb = openpyxl.load_workbook(file_path)
    except BadZipfile as e:
        # 次のファイルタイプのみ許容します。[.xlsx,.xlsm]
        status_code = '499-00401'
        raise AppException(status_code)     # noqa: F405
    
    # 対象メニューを特定するためのIDを取得する
    menu_id = menu_record[0].get('MENU_ID')
    menu_name = menu_record[0].get('MENU_NAME_' + lang.upper())
    
    # エラー判定用にシート名を取得
    sheets_name_list = wb.sheetnames
    
    # シート名が想定と違う場合はエラーとする
    if menu_name[:31] not in sheets_name_list:
        # このメニューの編集用Excelファイルではありません。
        status_code = '499-00402'
        raise AppException(status_code)     # noqa: F405
    
    ret = objdbca.table_select(t_common_menu_column_link, 'WHERE MENU_ID = %s AND DISUSE_FLAG = %s ORDER BY COLUMN_DISP_SEQ ASC', [menu_id, 0])
    
    # 項目順記憶用
    column_order = []
    # 登録更新時に除外する項目リスト
    register_list = []
    update_list = []
    column_class_file = ['9', '20']
    for recode in ret:
        column_name_rest = str(recode.get('COLUMN_NAME_REST'))
        column_class_id = str(recode.get('COLUMN_CLASS'))
        auto_input = str(recode.get('AUTO_INPUT'))
        input_item = str(recode.get('INPUT_ITEM'))
        view_item = str(recode.get('VIEW_ITEM'))
        
        # 登録更新時に不要な項目
        if auto_input == '1' or input_item == '0':
            register_list.append(column_name_rest)
        
        # カラムクラスIDがファイルアップロードのものは除外する
        if column_class_id in column_class_file:
            register_list.append(column_name_rest)
            update_list.append(column_name_rest)
        
        # Excelには表示しない項目
        if input_item == '2' and view_item == '0':
            continue
        
        # Excelには表示しない項目
        if column_class_id == '23':
            continue
        
        if column_name_rest == 'discard':
            # 廃止フラグは先頭に追加する
            column_order.insert(0, column_name_rest)
        else:
            column_order.append(column_name_rest)
    
    # 先頭のシートを選択する
    ws = wb.worksheets[0]
    
    # C列の実行処理種別を見て対象の行を記憶する
    row_num = ws.max_row
    col_num = ws.max_column
    # 登録\更新\廃止\復活
    msg_reg = g.appmsg.get_api_message('MSG-30004')
    msg_upd = g.appmsg.get_api_message('MSG-30005')
    msg_dis = g.appmsg.get_api_message('MSG-30006')
    msg_res = g.appmsg.get_api_message('MSG-30007')
    
    # エクセルに入力されたデータ
    excel_data = []
    # 実行処理種別
    process_type = []
    for i in range(row_num):
        # 対象行記憶用
        target_row = []
        # 実行処理種別を見る
        men_type = ws.cell(row=i + 1, column=3).value
        
        if men_type == msg_reg or men_type == msg_upd or men_type == msg_dis or men_type == msg_res:
            process_type.append(men_type)
            for j in range(4, col_num + 1):
                target_row.append(ws.cell(row=i + 1, column=j).value)
                
            excel_data.append(target_row)
    
    parameter = []
    dict_param = {}
    file_param = {}
    
    # i:対象行の分だけループ
    if excel_data:
        for row_i in range(len(excel_data)):
            dict_param = {}
            parameter.append({})
            parameter[row_i]["file"] = file_param
            # j:項目数の分だけループ
            for col_j in range(len(excel_data[0])):
                if process_type[row_i] == msg_reg:
                    if column_order[col_j] in register_list:
                        continue
                elif process_type[row_i] == msg_upd:
                    if column_order[col_j] in update_list:
                        continue
                dict_param[column_order[col_j]] = excel_data[row_i][col_j]
            parameter[row_i]["parameter"] = dict_param
            if process_type[row_i] == msg_reg:
                process_type[row_i] = "Register"
                parameter[row_i]["parameter"]["discard"] = "0"
            elif process_type[row_i] == msg_upd:
                process_type[row_i] = "Update"
                parameter[row_i]["parameter"]["discard"] = "0"
            elif process_type[row_i] == msg_dis:
                process_type[row_i] = "Update"
                parameter[row_i]["parameter"]["discard"] = "1"
            elif process_type[row_i] == msg_res:
                process_type[row_i] = "Update"
                parameter[row_i]["parameter"]["discard"] = "0"
                
            parameter[row_i]["type"] = process_type[row_i]
    
    # メニューのレコード登録/更新(更新/廃止/復活)
    result_data = menu_maintenance_all.rest_maintenance_all(objdbca, menu, parameter)
    
    # 処理が終わったらwbは削除する
    os.remove(file_path)
    
    return result_data
