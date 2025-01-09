#   Copyright 2022 NEC Corporation
#
#   Licensed under the Apache License, Version 2.0 (the "License");
#   you may not use this file except in compliance with the License.
#   You may obtain a copy of the License at
#
#       http://www.apache.org/licenses/LICENSE-2.0
#
#   Unless required by applicable law or agreed to in writing, software
#   distributed under the License is distributed on an "AS IS" BASIS,
#   WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
#   See the License for the specific language governing permissions and
#   limitations under the License.


from flask import g  # noqa: F401

import uuid  # noqa: F401
import textwrap
import sys
import traceback
from distutils.util import strtobool
import mimetypes
import base64
import re
import os
import random
import unicodedata
import shutil
import traceback

import difflib
from deepdiff import DeepDiff

import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles.fonts import Font
from openpyxl.utils.escape import *  # noqa: F403

from common_libs.common import *  # noqa: F403
from common_libs.loadtable import *  # noqa: F403
from common_libs.column import *  # noqa: F403
from common_libs.common import storage_access
from common_libs.common.util import get_iso_datetime, arrange_stacktrace_format


# 比較実行画面用情報取得(リスト情報、パラメータフォーマット)
def get_compares_data(objdbca, menu):
    """
        get compares data
        ARGS:
        RETRUN:
            {
                "list":{"compare":[],"host":[]},
                "dict":{"compare":{},"host":{}},
                "parameter_base_format":{}
                "file_compare_parameter_format":{}
            }
    """

    # get base parameter fromat
    compare_parameter_format, file_compare_parameter_format = _get_base_parameter()

    # set result
    result = {}
    result.setdefault("list", {})
    result.setdefault("dict", {})
    result.setdefault("execute_parameter_format", {})
    result["execute_parameter_format"].setdefault("compare", compare_parameter_format)
    result["execute_parameter_format"].setdefault("file", compare_parameter_format)
    result["list"].setdefault("compare", [])
    result["dict"].setdefault("compare", {})

    # get compare , menu
    table_name = "T_COMPARE_CONFG_LIST"
    where_str = "WHERE DISUSE_FLAG <> 1 ORDER BY COMPARE_NAME ASC "
    bind_list = []

    rows = objdbca.table_select(table_name, where_str, bind_list)

    taget_menu_ids = []
    for row in rows:
        tmp_id = row.get("COMPARE_ID")
        tmp_name = row.get("COMPARE_NAME")
        result["list"]["compare"].append(tmp_name)
        result["dict"]["compare"].setdefault(tmp_id, tmp_name)
        menu_id_1 = row.get("TARGET_MENU_1")
        menu_id_2 = row.get("TARGET_MENU_2")
        taget_menu_ids.append(menu_id_1)
        taget_menu_ids.append(menu_id_2)

    taget_menu_ids = list(set(taget_menu_ids))

    # get table_name
    sql_str = ""
    bind_list = []
    list_sql = []
    for taget_menu_id in taget_menu_ids:
        tmp_sql_str = " SELECT * FROM `T_COMN_MENU_TABLE_LINK` WHERE `DISUSE_FLAG` <> 1 AND `MENU_ID` = %s "
        list_sql.append(tmp_sql_str)
        bind_list.append(taget_menu_id)

    sql_str = " UNION ".join([str(tmp_sql) for tmp_sql in list_sql])

    host_ids = []
    result["list"].setdefault("host", [])
    result["dict"].setdefault("host", {})
    if sql_str:
        rows = objdbca.sql_execute(sql_str, bind_list)
        table_names = []
        for row in rows:
            table_name = row.get("TABLE_NAME")
            table_names.append(table_name)

        # get host list
        sql_str = ""
        bind_list = []
        list_sql = []
        for table_name in table_names:
            tmp_sql_str = textwrap.dedent("""
                SELECT
                    `TAB_A`.`HOST_ID`,
                    `TAB_B`.`HOST_NAME`
                FROM
                    `{table_name}` `TAB_A`
                LEFT JOIN `T_ANSC_DEVICE` `TAB_B` ON ( `TAB_A`.`HOST_ID` = `TAB_B`.`SYSTEM_ID` )
                WHERE `TAB_A`.`DISUSE_FLAG` <> 1
                AND `TAB_B`.`DISUSE_FLAG` <> 1
            """).format(table_name=table_name).strip()
            list_sql.append(tmp_sql_str)

        sql_str = " UNION ".join([str(tmp_sql) for tmp_sql in list_sql])

        # set result host data
        rows = objdbca.sql_execute(sql_str, bind_list)
        for row in rows:
            tmp_id = row.get("HOST_ID")
            tmp_name = row.get("HOST_NAME")
            host_ids.append(tmp_name)
            result["dict"]["host"].setdefault(tmp_id, tmp_name)

    host_ids = list(set(host_ids))
    host_ids.sort()
    result["list"]["host"] = host_ids

    return result


# 比較実行
def compare_execute(objdbca, menu, parameter, options={}, file_required=False):
    """
        compare_execute
        ARGS:
            parameter = {
                "compare": "",
                "base_date_1": "",
                "base_date_2": "",
                "host": [],
                "other_options": {}
            }
            options={}
        RETRUN:
            {"config":{},"compare_data":{},"compare_diff_flg":{}}
    """

    try:

        # set base config
        compare_config = _set_base_config(parameter)

        # parameter chk
        retBool, compare_config = _chk_parameters(objdbca, compare_config, options)
        if retBool is False:
            tmp_msg = _get_msg(compare_config, "parameter_error")
            status_code = "499-01001"
            log_msg_args = [tmp_msg]
            api_msg_args = [tmp_msg]
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

        # set compare mode , parameter format
        compare_config = _set_compare_mode(compare_config, options)

        # set input order lang
        compare_config = _set_input_order_lang(compare_config, options)

        # get compare config
        retBool, compare_config = _set_compare_config_info(objdbca, compare_config, options)
        if retBool is False:
            tmp_msg = _get_msg(compare_config, "compare_config")
            status_code = "499-01002"
            log_msg_args = [tmp_msg]
            api_msg_args = [tmp_msg]
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

        # get data
        retBool, compare_config = _get_target_datas(objdbca, compare_config, options)
        if retBool is False:
            tmp_msg = _get_msg(compare_config, "compare_target_menu_error")
            status_code = "499-01004"
            log_msg_args = [tmp_msg]
            api_msg_args = [tmp_msg]
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

        # compare data execute
        retBool, compare_config = _execute_compare_data(objdbca, compare_config, options, file_required)
        if retBool is False:
            tmp_msg = _get_msg(compare_config, "execute_compare")
            status_code = "499-01005"
            log_msg_args = [tmp_msg]
            api_msg_args = [tmp_msg]
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

        compare_mode = _get_flg(compare_config, "compare_mode")
        output_flg = _get_flg(compare_config, "output_flg")

        # override_column_info_vertical_ver  item1 -> item[input_order:1]
        vertical_compare_flg = _get_flg(compare_config, "vertical_compare_flg")
        if vertical_compare_flg is True:
            retBool, compare_config = _override_column_info_vertical_ver(objdbca, compare_config, options)

        # set return
        result = {}
        result.setdefault("config", {})
        if compare_mode == "normal":
            result.setdefault("compare_data", {})
            result.setdefault("compare_diff_flg", {})
            # set result->config
            result["config"].setdefault("target_menus_rest", list(compare_config.get("target_menus").values()))
            result["config"].setdefault("target_menus", list(compare_config.get("target_menus_lang").values()))
            result["config"].setdefault("target_host_list", compare_config.get("target_host_list"))
            result["config"].setdefault("target_column_info", compare_config.get("column_info"))
            result["config"].setdefault("compare_file", _get_flg(compare_config, "compare_file"))
            # set result->compare_diff_flg
            result["compare_diff_flg"] = compare_config.get("result_compare_host")
            # set result->compare_data
            result["compare_data"] = compare_config.get("compare_data_ptn1")

        elif compare_mode == "file":
            # set result->config
            result["config"].setdefault("target_compare_file_info", compare_config.get("target_compare_file_info"))
            result["config"].setdefault("compare_file", _get_flg(compare_config, "compare_file"))
            # set result->unified_diff
            result.setdefault("unified_diff", {})
            result["unified_diff"] = compare_config.get("unified_diff")

        if output_flg is True:
            # excel output
            result = _create_outputfile(objdbca, compare_config, result, options)

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "execute_compare_error", True)
        retBool = False

    return result


# 比較パラメータフォーマット取得
def _get_base_parameter():
    """
        get base parameter
        ARGS:
        RETRUN:
            compare_parameter_format, file_compare_parameter_format
    """
    # base format compare / file compare
    compare_parameter_format = {
        "compare": "",
        "base_date_1": "",
        "base_date_2": "",
        "host": [],
        # "other_options": {}
    }
    file_compare_parameter_format = {
        "compare": "",
        "base_date_1": "",
        "base_date_2": "",
        "host": "",
        "copmare_target_column": "",
        # "other_options": {}
    }
    return compare_parameter_format, file_compare_parameter_format


# config初期設定
def _set_base_config(parameter):
    """
        set base config [caller:compare_execute]
        ARGS:
            parameter
            mode: normal, file
        RETRUN:
            {}
    """

    # accept file mimetype
    accept_compare_file_list = [
        "text/plain",
        "application/json",
        "text/csv",
        "text/html",
        "text/javascript",
        "application/ld+json",
        'text/x-yaml',
        'text/yaml',
        'application/yaml',
        'application/x-yaml',
    ]

    add_ext_mimetype = {
        ".yaml": [
            "text/yaml",
            'text/x-yaml',
            'text/yaml',
            'application/yaml',
            'application/x-yaml',
        ],
        ".yml": [
            "text/yaml",
            'text/x-yaml',
            'text/yaml',
            'application/yaml',
            'application/x-yaml',
        ]
    }

    for k_ext, l_mimes in add_ext_mimetype.items():
        for v_mime in l_mimes:
            mimetypes.add_type(v_mime, k_ext, False)

    ban_ext = [
        "zip",
        "gzip",
        "gz",
        "tar",
        "rar",
        "7z",
        "bzip2",
        "bz2",
        "exe",
        "lzh",
    ]

    accept_compare_file_list.extend([_v for _k, _v in mimetypes.types_map.items() if _v.startswith('text') or _v.startswith('application')])
    accept_compare_file_list.extend([_v for _k, _v in mimetypes.common_types.items() if _v.startswith('text') or _v.startswith('application')])
    accept_compare_file_list = list(set(accept_compare_file_list))

    for _v in accept_compare_file_list:
        _ind = accept_compare_file_list.index(_v)
        if _v is not None and (_v.startswith('text') or _v.startswith('application')):
            for _bk in ban_ext:
                if _v.find(_bk) != -1:
                    accept_compare_file_list.pop(_ind)

    # not use rest_key
    del_parameter_list = [
        # "uuid",
        # "host_name",
        "operation_name_select",
        # "operation_name_disp",
        # "base_datetime",
        "operation_date",
        "last_execute_timestamp",
        "remarks",
        "discard",
        "last_update_date_time",
        "last_updated_user",
    ]

    # not use rest_key
    no_compare_parameter_list = [
        "menu",
        "uuid",
        "host_name",
        "operation_name_select",
        "operation_name_disp",
        "base_datetime",
        "operation_date",
        "input_order",
        "last_execute_timestamp",
        "remarks",
        "discard",
        "last_update_date_time",
        "last_updated_user",
    ]

    # no_input_order_list
    no_input_order_list = [
        "menu",
        "uuid",
        "host_name",
        "operation_name_select",
        "operation_name_disp",
        "base_datetime",
        "operation_date",
        "input_order",
        "last_execute_timestamp",
        "remarks",
        "discard",
        "last_update_date_time",
        "last_updated_user",
    ]

    # not use rest_key
    vertical_no_use_rest_key = [
        "menu",
        "uuid",
        "host_name",
        "operation_name_select",
        "operation_name_disp",
        "base_datetime",
        "operation_date",
        "input_order",
        "last_execute_timestamp",
        "remarks",
        "discard",
        "last_update_date_time",
        "last_updated_user",
    ]

    input_order_langs = {
        "ja": "代入順序",
        "en:": "Input order"
    }

    compare_config = {}

    # parameter設定
    compare_config.setdefault("parameter", parameter)
    # LANGUAGE
    compare_config.setdefault("language", g.get('LANGUAGE'))
    # flg管理
    compare_config.setdefault("_flg_data", {})
    # 比較設定
    compare_config.setdefault("config_compare_list", {})
    # 比較設定詳細
    compare_config.setdefault("config_compare_detail", {})
    # 表示項目順LIST
    compare_config.setdefault("config_compare_disp", [])
    # 項目詳細(表示項目、紐付、カラムクラス等)
    compare_config.setdefault("column_info", {})
    # 比較元データ {"HOST":{"INPUT_ORDER":{"menu_1":{},"menu_2":{}}}}
    compare_config.setdefault("origin_data", {})
    # ホスト一覧(比較元データ)
    compare_config.setdefault("target_host_list", [])
    # 対象メニュー情報 {"menu_1":{"menu_name",,,},"menu_2":{"menu_name",,,}}
    compare_config.setdefault("target_menu_info", {})
    # 対象メニュー {"menu_1":"rest_name","menu_2":"rest_name"}
    compare_config.setdefault("target_menus", {"menu_1": "", "menu_2": ""})
    # 対象メニュー {"menu_1":"rest_name","menu_2":"rest_name"}
    compare_config.setdefault("target_menus_lang", {"menu_1": "", "menu_2": ""})
    # 代入順序リスト {"menu_1":[],"menu_2":[]}
    compare_config.setdefault("input_order_list", [])
    # 対象メニューobjtable {"menu_1": objtable, "menu_2": objtable}
    compare_config.setdefault("objtable", {})
    # 比較結果(ホスト毎)
    compare_config.setdefault("result_compare_host", {})
    # 比較対象データ無し時の項目補完用 {"menu_1":{"key_name":None},"menu_2":{"key_name":None}}
    compare_config.setdefault("base_parameter", {})
    compare_config.setdefault("base_parameter_all", {})
    # ファイル比較対象結果(unified_diff)
    compare_config.setdefault("unified_diff", {"file_data": {}, "diff_result": {}})
    # 不要項目リスト(rest_key)
    compare_config.setdefault("del_parameter_list", del_parameter_list)
    # 比較対象外項目リスト(rest_key+menu)
    compare_config.setdefault("no_compare_parameter_list", no_compare_parameter_list)
    # 縦型比較不要項目リスト
    compare_config.setdefault("vertical_no_use_rest_key", vertical_no_use_rest_key)
    # ファイル比較対象(mimetype)
    compare_config.setdefault("accept_compare_file_list", accept_compare_file_list)
    # 代入順序対象外リスト
    compare_config.setdefault("no_input_order_list", no_input_order_list)
    # 代入順序表示用言語
    compare_config.setdefault("input_order_langs", input_order_langs)
    # 代入順序付き項目用の代入順序
    compare_config.setdefault("input_order_lang", None)

    return compare_config


# 比較種別、対応フォーマット設定
def _set_compare_mode(compare_config, options):
    """
        set compare mode  [caller:compare_execute]
        ARGS:
            compare_config, flg_key, flg_val
        RETRUN:
            compare_config
    """
    # get parameter
    parameter = compare_config.get("parameter")
    # get base parameter format
    compare_parameter_format, file_compare_parameter_format = _get_base_parameter()

    # compare mode set parameter format marge
    compare_mode = options.get("compare_mode")
    if compare_mode == "normal":
        base_parameter = compare_parameter_format
    elif compare_mode == "file":
        base_parameter = file_compare_parameter_format
        tmp_parameter = compare_config.get("parameter")
        copmare_target_column = tmp_parameter.get("copmare_target_column")
        compare_config["copmare_target_column"] = copmare_target_column
    else:
        compare_mode = "normal"
        base_parameter = compare_parameter_format
    base_parameter.update(parameter)

    output_flg = options.get("output_flg")

    # set compare mode / output flg
    compare_config = _set_flg(compare_config, "compare_mode", compare_mode)
    compare_config = _set_flg(compare_config, "output_flg", output_flg)

    # set parameter
    compare_config["parameter"] = base_parameter

    return compare_config


# 代入順序表示名設定
def _set_input_order_lang(compare_config, options):
    """
        set input order lang  [caller:compare_execute]
        ARGS:
            compare_config, options
        RETRUN:
            compare_config
    """
    # get language
    language = compare_config.get("language")

    # get input_order_lang
    input_order_langs = compare_config.get("input_order_langs")
    input_order_lang = input_order_langs.get(language)

    input_order_lang = None

    # set input_order_lang
    compare_config["input_order_lang"] = input_order_lang

    return compare_config


# flg管理 設定 _flg_data
def _set_flg(compare_config, flg_key, flg_val):
    """
        set flg
        ARGS:
            compare_config, flg_key, flg_val
        RETRUN:
            compare_config
    """
    compare_config.setdefault("_flg_data", {})
    if flg_key not in compare_config["_flg_data"]:
        compare_config["_flg_data"].setdefault(flg_key, flg_val)
    else:
        compare_config["_flg_data"][flg_key] = flg_val
    return compare_config


# flg管理 取得 _flg_data
def _get_flg(compare_config, flg_key):
    """
        get flg
        ARGS:
            compare_config, flg_key
        RETRUN:
            flg_val
    """
    flg_val = None
    compare_config.setdefault("_flg_data", {})
    if flg_key in compare_config["_flg_data"]:
        flg_val = compare_config["_flg_data"][flg_key]
    return flg_val


# msg管理 設定 _msg
def _set_msg(compare_config, key, msg):
    """
        set msg
        ARGS:
            compare_config, key, msg
        RETRUN:
            compare_config
    """
    compare_config.setdefault("_msg", {})
    if key not in compare_config["_msg"]:
        compare_config["_msg"].setdefault(key, [])
    compare_config["_msg"][key].append(msg)
    return compare_config


# msg管理 取得 _msg
def _get_msg(compare_config, key):
    """
        get msg
        ARGS:
            compare_config, key
        RETRUN:
            msg
    """
    msg = []
    compare_config.setdefault("_msg", {})
    if key in compare_config["_msg"]:
        msg = compare_config["_msg"][key]
    return msg


# 比較実行処理
def _execute_compare_data(objdbca, compare_config, options, file_required=False):
    """
        execute compare data  [caller:compare_execute]
        ARGS:
            objdbca, compare_config, options
        RETRUN:
            {}
    """
    retBool = True

    try:
        # get column_info
        column_info = compare_config.get("column_info")
        # get origin_data
        origin_data = compare_config.get("origin_data")
        # get compare_mode  normal / file
        compare_mode = _get_flg(compare_config, "compare_mode")
        # get copmare_target_column
        copmare_target_column = compare_config.get("copmare_target_column")
        target_menus = compare_config.get("target_menus_lang")
        # get vertical_compare_flg
        vertical_compare_flg = _get_flg(compare_config, "vertical_compare_flg")

        input_order_lang = compare_config.get("input_order_lang")

        result_ptn1 = {}

        menu_name_1 = target_menus.get("menu_1")
        menu_name_2 = target_menus.get("menu_2")

        no_compare_target = []
        no_view_target_list = {}
        # for target host
        for target_host in list(origin_data.keys()):

            target_data_1 = {}
            target_data_2 = {}
            diff_flg_data = {}
            diff_flg_file = {}
            file_compare_info = {}
            no_view_target_list[target_host] = []
            origin_data[target_host].setdefault("__no_input_order__", {"menu_1": {}, "menu_2": {}})
            origin_data[target_host]["__no_input_order__"]["menu_1"].setdefault("menu", menu_name_1)
            origin_data[target_host]["__no_input_order__"]["menu_2"].setdefault("menu", menu_name_2)

            # target col
            for tmp_info in column_info:
                file_mimetypes = {}
                other_options = {}
                tmp_file_compare_info = {}

                # set target data
                menu_1_data = {}
                target_uuid_1 = None
                col_val_menu_1 = None
                base_datetime_1 = None
                operation_name_disp_1 = None
                menu_2_data = {}
                target_uuid_2 = None
                col_val_menu_2 = None
                base_datetime_2 = None
                operation_name_disp_2 = None

                # get col name
                col_name = tmp_info.get("col_name")

                # compare target flg
                compare_target_flg = tmp_info.get("compare_target_flg")
                # compare file flg
                file_flg = tmp_info.get("file_flg")
                # no_input_order_flg
                no_input_order_flg = tmp_info.get("no_input_order_flg")

                # menu rest_name
                col_name_menu_1 = tmp_info.get("menu_1")
                col_name_menu_2 = tmp_info.get("menu_2")
                # menu columnclass
                column_class_name_1 = tmp_info.get("column_class_1")
                column_class_name_2 = tmp_info.get("column_class_2")

                # check target column
                target_column_flg = _chk_target_column_flg(copmare_target_column, col_name)

                # no compare target list
                if compare_target_flg is False:
                    no_compare_target.append(col_name)

                # target vertical menu
                if vertical_compare_flg is True:
                    input_orders = list(origin_data[target_host].keys())
                    input_orders = list(set(input_orders))
                    # get target input order :column item_1[x] -> x
                    try:
                        copmare_target_column_io = re.findall(r'\[.*\]', copmare_target_column)
                        copmare_target_column_io = copmare_target_column_io[0].replace('[', '').replace(']', '')
                        input_orders = [int(copmare_target_column_io)]
                        target_column_flg = True
                    except Exception:
                        t = traceback.format_exc()
                        g.applogger.info("[timestamp={}] {}".format(str(get_iso_datetime()), arrange_stacktrace_format(t)))
                        copmare_target_column_io = None
                else:
                    input_orders = ["__no_input_order__"]

                # target filter
                if target_column_flg is True:
                    for tmp_order in input_orders:
                        file_mimetypes = {}
                        other_options = {}
                        tmp_file_compare_info = {}

                        input_order_1 = tmp_order
                        input_order_2 = tmp_order
                        tmp_col_name = "{}".format(col_name)

                        # get target menu data -> uuid value base_datetime operation_name_disp
                        menu_1_data, target_uuid_1, col_val_menu_1, base_datetime_1, operation_name_disp_1 = \
                            _get_line_values("menu_1", origin_data, target_host, input_order_1, col_name_menu_1)
                        menu_2_data, target_uuid_2, col_val_menu_2, base_datetime_2, operation_name_disp_2 = \
                            _get_line_values("menu_2", origin_data, target_host, input_order_2, col_name_menu_2)

                        # set menu_name
                        if menu_name_1 == col_val_menu_1:
                            if tmp_col_name in target_data_1:
                                target_data_1[tmp_col_name] = menu_name_1
                        if menu_name_2 == col_val_menu_2:
                            if tmp_col_name in target_data_2:
                                target_data_2[tmp_col_name] = menu_name_2

                        # vertical menu and comare target parameters
                        tmp_col_name = _override_col_name_io(
                            col_name,
                            tmp_col_name,
                            tmp_order,
                            input_order_lang,
                            no_input_order_flg,
                            vertical_compare_flg
                        )

                        # set no use target list
                        none_column_class_list = [
                            "PasswordColumn"
                            "SensitiveSingleTextColumn",
                            "SensitiveMultiTextColumn",
                            "PasswordIDColumn",
                            "JsonPasswordIDColumn",
                        ]
                        if col_val_menu_1 is None and col_val_menu_2 is None:
                            if column_class_name_1 not in none_column_class_list and \
                                    column_class_name_2 not in none_column_class_list:
                                no_view_target_list[target_host].append(tmp_col_name)

                        # check target colname
                        if tmp_col_name is not None:
                            # set target key->value
                            target_data_1.setdefault(tmp_col_name, col_val_menu_1)
                            target_data_2.setdefault(tmp_col_name, col_val_menu_2)

                            # get file data
                            if file_flg is True:
                                # get file compare resut: file_compare_info, diff_flg_file
                                compare_config, tmp_file_compare_info, diff_flg_file = _get_compare_file_result(
                                    objdbca,
                                    compare_config,
                                    diff_flg_file,
                                    file_mimetypes,
                                    target_host,
                                    tmp_col_name,
                                    compare_target_flg,
                                    file_required,
                                    dict_menu1={
                                        "col_name_menu": col_name_menu_1,
                                        "target_uuid": target_uuid_1,
                                        "col_val_menu": col_val_menu_1,
                                        "column_class_name": column_class_name_1,
                                        "operation_name_disp": operation_name_disp_1,
                                        "base_datetime": base_datetime_1,
                                    },
                                    dict_menu2={
                                        "col_name_menu": col_name_menu_2,
                                        "target_uuid": target_uuid_2,
                                        "col_val_menu": col_val_menu_2,
                                        "column_class_name": column_class_name_2,
                                        "operation_name_disp": operation_name_disp_2,
                                        "base_datetime": base_datetime_2,
                                    }
                                )
                            else:
                                diff_flg_file.setdefault(tmp_col_name, False)

                            # set tmp_file_compare_info
                            file_compare_info.setdefault(tmp_col_name, tmp_file_compare_info)
                            other_options.setdefault("file_compare_info", tmp_file_compare_info)

            if compare_mode == "normal":
                negative_key = [
                    g.appmsg.get_api_message('MSG-60028'),  # 'メニュー名', 'menu',
                    g.appmsg.get_api_message('MSG-60029'),  # '項番', 'uuid',
                    f"{g.appmsg.get_api_message('MSG-90046')}/{g.appmsg.get_api_message('MSG-60030')}", # 'オペレーション名', 'operation_name_disp',
                    f"{g.appmsg.get_api_message('MSG-90046')}/{g.appmsg.get_api_message('MSG-60031')}"  # '基準日時', 'base_datetime',
                ]
                # target only
                tmp_target_data_1 = {k: v for k, v in target_data_1.items() if k not in negative_key}
                tmp_target_data_2 = {k: v for k, v in target_data_2.items() if k not in negative_key}

                # diff 全体
                result_deepdiff = DeepDiff(tmp_target_data_1, tmp_target_data_2)
                tmp_compare_result = False
                diff_keys_set = set()
                for diff_type in ["type_changes","dictionary_item_added", "dictionary_item_removed", "values_changed"]:
                    if diff_type in result_deepdiff:
                        for key in result_deepdiff[diff_type]:
                            diff_keys_set.add(key[6:-2])
                diff_key = list(diff_keys_set)

                if len(result_deepdiff) != 0:
                    tmp_compare_result = True
                else:
                    pass
                compare_config["result_compare_host"].setdefault(target_host, tmp_compare_result)


                # get target rest_key
                target_data_key = list(target_data_1.keys())
                target_data_key.extend(list(target_data_2.keys()))

                # get file target list
                file_list_colneme = [tmp_info.get('col_name') for tmp_info in column_info if tmp_info.get("file_flg") is True]

                # set diff flg[value]
                for tmp_key in target_data_key:
                    diff_flg = False
                    if tmp_key in diff_key and tmp_key not in no_compare_target:
                        diff_flg = True
                    if vertical_compare_flg is True and tmp_key in no_view_target_list[target_host]:
                        diff_flg = None

                    if tmp_key in file_list_colneme:
                        file_diff_flg = diff_flg_file.get(tmp_key)
                        if diff_flg is False and file_diff_flg is True:
                            # override file_diff_flg->result_compare_host
                            compare_config["result_compare_host"][target_host] = True

                    diff_flg_data.setdefault(tmp_key, diff_flg)

                result_ptn1.setdefault(target_host, {})
                result_ptn1[target_host].setdefault("target_data_1", target_data_1)
                result_ptn1[target_host].setdefault("target_data_2", target_data_2)
                result_ptn1[target_host].setdefault("compare_diff_flg", tmp_compare_result)
                result_ptn1[target_host].setdefault("_data_diff_flg", diff_flg_data)
                result_ptn1[target_host].setdefault("_file_compare_execute_flg", diff_flg_file)
                result_ptn1[target_host].setdefault("_file_compare_execute_info", file_compare_info)
                compare_config["compare_data_ptn1"] = result_ptn1

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "execute_compare_error", True)
        retBool = False
    return retBool, compare_config,


# 比較対象項目判定
def _chk_target_column_flg(copmare_target_column, col_name):
    """
        check compare target
        ARGS:
            copmare_target_column, col_name,
        RETRUN:
            bool
    """
    target_column_flg = True
    if copmare_target_column is None:
        pass
    elif col_name == copmare_target_column:
        pass
    else:
        target_column_flg = False
    return target_column_flg


# 値の比較判定
def _chk_value_compare_flg(compare_target_flg, col_val_menu_1, col_val_menu_2):
    """
        check compare value
        ARGS:
            compare_target_flg, col_val_menu_1, col_val_menu_2
        RETRUN:
            bool
    """
    value_compare_flg = False
    if compare_target_flg is True:
        if col_val_menu_1 != col_val_menu_2:
            value_compare_flg = True
        else:
            value_compare_flg = False
    else:
        value_compare_flg = False
    return value_compare_flg


# 比較対象関連の情報取得[対象行,ID,値,基準日,オペレーション]
def _get_line_values(target_key, origin_data, target_host, input_order, col_name_menu):
    """
        get compare data for target menu line
        ARGS:
            target_key, origin_data, target_host, input_order, col_name_menu
        RETRUN:
            menu_data, target_uuid, col_val_menu, base_datetime, operation_name_disp
    """
    menu_data = {}
    target_uuid = None
    col_val_menu = None
    base_datetime = None
    operation_name_disp = None
    try:
        menu_data = origin_data[target_host][input_order].get(target_key)
        target_uuid = menu_data.get("uuid")
        # col_val_menu->str()[:is not None]
        col_val_menu = str(menu_data.get(col_name_menu)) if menu_data.get(col_name_menu) is not None else menu_data.get(col_name_menu)
        base_datetime = menu_data.get("base_datetime")
        operation_name_disp = menu_data.get("operation_name_disp")
    except Exception:
        t = traceback.format_exc()
        g.applogger.info("[timestamp={}] {}".format(str(get_iso_datetime()), arrange_stacktrace_format(t)))
        pass
    return menu_data, target_uuid, col_val_menu, base_datetime, operation_name_disp,


# ファイル比較結果(string:unified形式)
def _get_unified_diff(accept_compare_file_list, filename_1, filename_2, mimetype_1, mimetype_2, data_1, data_2, file_required=False):
    """
        get unified diff data: format string
        ARGS:
            accept_compare_file_list, filename_1, filename_2, mimetype_1, mimetype_2, data_1, data_2
        RETRUN:
            str_diff
    """
    filename_1 = "" if filename_1 is None or len(filename_1) == 0 else filename_1
    filename_2 = "" if filename_2 is None or len(filename_2) == 0 else filename_2
    chunk_byte = 10000

    file_diff = False
    str_rdiff = ""
    try:
        if file_required is True:
            data_1_linebreak = base64.b64decode(data_1.encode()).decode().splitlines()
            data_2_linebreak = base64.b64decode(data_2.encode()).decode().splitlines()
        else:
            data_1_linebreak = []
            data_2_linebreak = []
            if data_1:
                with open(data_1, "rb") as tmp_data_1:
                    while chunk_1:= tmp_data_1.read(chunk_byte):
                        data_1_linebreak.extend(chunk_1.decode().splitlines())
            if data_2:
                with open(data_2, "rb") as tmp_data_2:
                    while chunk_2:= tmp_data_2.read(chunk_byte):
                        data_2_linebreak.extend(chunk_2.decode().splitlines())
    except (UnicodeDecodeError, ValueError):
        if not data_1 and not data_2:
            file_diff = False
        elif not data_1 or not data_2:
            file_diff = True
        else:
            with open(data_1, "rb") as tmp_data_1, open(data_2, "rb") as tmp_data_2:
                while chunk_1 := tmp_data_1.read(chunk_byte):
                    chunk_2 = tmp_data_2.read(chunk_byte)
                    if chunk_1 != chunk_2:
                        file_diff = True
        return file_diff, str_rdiff
    except Exception:
        t = traceback.format_exc()
        g.applogger.info("[timestamp={}] {}".format(str(get_iso_datetime()), arrange_stacktrace_format(t)))
        # read file is faild
        status_code = "499-01006"
        log_msg_args = [filename_1, filename_2]
        api_msg_args = [filename_1, filename_2]
        raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    rdiff = difflib.unified_diff(
        data_1_linebreak,
        data_2_linebreak,
        filename_1,
        filename_2,
        lineterm='')
    str_rdiff = '\n'.join(rdiff)

    if str_rdiff:
        file_diff = True
    return file_diff, str_rdiff


# 縦型メニュー項目名:代入順序入り変換上書き
def _override_col_name_io(col_name, col_name_io, input_order, input_order_lang, no_input_order_flg, vertical_compare_flg):
    """
        override col_name add input_order
        ARGS:
            col_name, col_name_io, input_order, no_input_order_flg, vertical_compare_flg
        RETRUN:
            col_name_io
    """
    if input_order != "__no_input_order__" and no_input_order_flg is False:
        col_name_io = "{}[{}]".format(col_name, input_order)
        col_name_io = _get_col_name_input_order(col_name, input_order, input_order_lang)
    else:
        # vertical menu no use target colname: tmp_col_name->None
        if no_input_order_flg is False and vertical_compare_flg is True:
            col_name_io = None
    return col_name_io


# ファイル比較結果のフラグ、詳細
def _get_compare_file_result(objdbca, compare_config, diff_flg_file, file_mimetypes, target_host, tmp_col_name, target_flg, file_required, dict_menu1, dict_menu2):
    """
        get compare file file_compare_info and set diff_flg_file
        ARGS:
            objdbca, compare_config, diff_flg_file, file_mimetypes, target_host, tmp_col_name, target_flg, dict_menu1, dict_menu2
        RETRUN:
            compare_config, tmp_file_compare_info, diff_flg_file
    """
    # objtable
    objtable_1 = compare_config["objtable"]["menu_1"]
    objtable_2 = compare_config["objtable"]["menu_2"]

    compare_mode = _get_flg(compare_config, "compare_mode")
    accept_compare_file_list = compare_config.get("accept_compare_file_list")
    tmp_file_data_1 = None
    tmp_file_data_2 = None
    tmp_file_mimetype_1 = None
    tmp_file_mimetype_2 = None
    col_val_menu_1 = dict_menu1.get("col_val_menu")
    col_name_menu_1 = dict_menu1.get("col_name_menu")
    target_uuid_1 = dict_menu1.get("target_uuid")
    column_class_name_1 = dict_menu1.get("column_class_name")
    operation_name_disp_1 = dict_menu1.get("operation_name_disp")
    base_datetime_1 = dict_menu1.get("base_datetime")

    col_val_menu_2 = dict_menu2.get("col_val_menu")
    col_name_menu_2 = dict_menu2.get("col_name_menu")
    target_uuid_2 = dict_menu2.get("target_uuid")
    column_class_name_2 = dict_menu2.get("column_class_name")
    operation_name_disp_2 = dict_menu2.get("operation_name_disp")
    base_datetime_2 = dict_menu2.get("base_datetime")

    file_mimetypes.setdefault(tmp_col_name, {"target_data_1": {}, "target_data_2": {}})

    # compare result[value]
    value_compare_flg = _chk_value_compare_flg(target_flg, col_val_menu_1, col_val_menu_2)

    # get file base64 string  file mimetype
    if col_val_menu_1 is not None:
        tmp_file_data_1, tmp_file_mimetype_1 = _get_file_data_columnclass(
            objdbca,
            objtable_1,
            col_name_menu_1,
            col_val_menu_1,
            target_uuid_1,
            column_class_name_1,
            file_required)
        file_mimetypes[tmp_col_name]["target_data_1"].setdefault(col_val_menu_1, tmp_file_mimetype_1)
    else:
        col_val_menu_1 = "no file"
        tmp_file_data_1 = ""
        tmp_file_mimetype_1 = "text/plain"
    if col_val_menu_2 is not None:
        tmp_file_data_2, tmp_file_mimetype_2 = _get_file_data_columnclass(
            objdbca,
            objtable_2,
            col_name_menu_2,
            col_val_menu_2,
            target_uuid_2,
            column_class_name_2,
            file_required)
        file_mimetypes[tmp_col_name]["target_data_2"].setdefault(col_val_menu_2, tmp_file_mimetype_2)
    else:
        col_val_menu_2 = "no file"
        tmp_file_data_2 = ""
        tmp_file_mimetype_2 = "text/plain"
    # compare result[file]

    # compare file
    if not file_required and (os.path.exists(tmp_file_data_1) or os.path.exists(tmp_file_data_2)):
        file_diff, str_rdiff = _get_unified_diff(
            accept_compare_file_list,
            col_val_menu_1,
            col_val_menu_2,
            tmp_file_mimetype_1,
            tmp_file_mimetype_2,
            tmp_file_data_1, # file_required=Falseの場合、fileのパスが入る
            tmp_file_data_2, # file_required=Falseの場合、fileのパスが入る
            file_required
        )
        if file_diff and str_rdiff:
            file_compare_flg = True
            diff_flg_file.setdefault(tmp_col_name, file_compare_flg)
        elif file_diff:
            value_compare_flg = False
            file_compare_flg = True
            diff_flg_file.setdefault(tmp_col_name, file_compare_flg)
        else:
            file_compare_flg = False
            diff_flg_file.setdefault(tmp_col_name, file_compare_flg)
    else:
        if col_val_menu_1 != col_val_menu_2:
            value_compare_flg = False
            file_compare_flg = True
            diff_flg_file.setdefault(tmp_col_name, file_compare_flg)
        else:
            file_compare_flg = False
            diff_flg_file.setdefault(tmp_col_name, file_compare_flg)

    if compare_mode == "file":

        if str_rdiff == "":
            compare_config = _set_flg(compare_config, "compare_file", False)

        compare_config["unified_diff"]["file_data"].setdefault(
            "menu_1",
            {"name": col_val_menu_1, "data": "" if file_required is False else tmp_file_data_1}
        )
        compare_config["unified_diff"]["file_data"].setdefault(
            "menu_2",
            {"name": col_val_menu_2, "data": "" if file_required is False else tmp_file_data_2}
        )
        compare_config["unified_diff"]["diff_result"] = str_rdiff
        target_compare_file_info = {}
        target_compare_file_info.setdefault("target_host", target_host)
        target_compare_file_info.setdefault("operation_1", operation_name_disp_1)
        target_compare_file_info.setdefault("base_datetime_1", base_datetime_1)
        target_compare_file_info.setdefault("operation_2", operation_name_disp_2)
        target_compare_file_info.setdefault("base_datetime_2", base_datetime_2)

        compare_config["target_compare_file_info"] = target_compare_file_info

    # set compare file endpoin + parameter
    endpoint = ("/api/{organization_id}/workspaces/{workspace_id}/ita/menu/{menu}/compare/execute/file?file=no".format(
                organization_id=g.get("ORGANIZATION_ID"),
                workspace_id=g.get("WORKSPACE_ID"),
                menu="compare_execute",
                ))
    file_compare_parameter = compare_config.get("parameter").copy()
    file_compare_parameter.setdefault("copmare_file_data", True)
    file_compare_parameter.setdefault("copmare_target_column", tmp_col_name)
    # file_compare_parameter.setdefault("host", target_host)
    file_compare_parameter["host"] = target_host

    # set compare result detail[file]
    tmp_file_compare_info = {}
    tmp_file_compare_info["file_name_diff_flg"] = value_compare_flg
    tmp_file_compare_info["file_data_diff_flg"] = file_compare_flg
    tmp_file_compare_info.setdefault("file_mimetype", file_mimetypes)
    tmp_file_compare_info.setdefault("file_name_diff_flg", value_compare_flg)
    tmp_file_compare_info.setdefault("file_data_diff_flg", file_compare_flg)
    tmp_file_compare_info.setdefault("method", "POST")
    tmp_file_compare_info.setdefault("endpoint", endpoint)
    tmp_file_compare_info.setdefault("parameter", file_compare_parameter)

    del tmp_file_data_1
    del tmp_file_data_2

    return compare_config, tmp_file_compare_info, diff_flg_file


# 対象メニューデータ取得
def _get_target_datas(objdbca, compare_config, options):
    """
        get target datas  [caller:compare_execute]
        ARGS:
            objdbca, compare_config, options
        RETRUN:
            {}
    """
    retBool = True

    try:
        target_parameter = {}
        target_host_list = []
        target_menus = compare_config.get("target_menus")
        tmp_accept_host_list = compare_config.get("parameter").get("host")
        if isinstance(tmp_accept_host_list, list):
            accept_host_list = tmp_accept_host_list
        elif isinstance(tmp_accept_host_list, str):
            accept_host_list = [tmp_accept_host_list]
        else:
            accept_host_list = []

        target_menu_info = compare_config.get("target_menu_info")

        input_order_list = {}
        host_filter_flg = False
        if len(accept_host_list) != 0:
            host_filter_flg = True
        compare_config = _set_flg(compare_config, "host_filter_flg", host_filter_flg)

        for target_key, target_menu in target_menus.items():
            # load load_table
            try:
                objmenu = load_table.loadTable(objdbca, target_menu)  # noqa: F405
            except Exception:
                t = traceback.format_exc()
                g.applogger.info("[timestamp={}] {}".format(str(get_iso_datetime()), arrange_stacktrace_format(t)))
                # loadTable is faild
                status_code = "401-00003"
                log_msg_args = [target_menu]
                api_msg_args = [target_menu]
                raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

            # set objtable base_parameters
            restkey_list = objmenu.get_restkey_list()
            compare_config["objtable"].setdefault(target_key, objmenu.get_objtable())
            compare_config = _set_base_parameters(compare_config, restkey_list, target_key)

            input_order_list.setdefault(target_key, {})
            vertical = target_menu_info.get(target_key).get("vertical")

            # set filter parameter
            rest_mode = "excel"
            filter_parameter = {"discard": {"NORMAL": "0"}}
            operation_date = None
            # set base_date
            if target_key == "menu_1":
                operation_date = compare_config.get("parameter").get("base_date_1")
            if target_key == "menu_2":
                operation_date = compare_config.get("parameter").get("base_date_2")

            if operation_date:
                operation_date_add = operation_date + ".999999" # issue2445 リクエストにマイクロ秒がないために検索対象からもれてしまうケースがあるため、リクエストの秒代全てがヒットするようにする
                filter_parameter.setdefault("base_datetime", {"RANGE": {"END": operation_date_add}})

            # set host
            if host_filter_flg is True:
                filter_parameter.setdefault("host_name", {"LIST": accept_host_list})
            # execute filter
            status_code, tmp_result, msg = objmenu.rest_filter(filter_parameter, rest_mode)

            # del discard operation data
            tmp_result = [_tmp_data for _tmp_data in tmp_result if _tmp_data['parameter']['base_datetime'] is not None]

            # target_parameter SET
            compare_config.setdefault("_target_parameter", {})
            compare_config = _set_target_parameter(compare_config, target_key, vertical, tmp_result)
            target_parameter = compare_config.get("_target_parameter")

            # get target host list
            if len(target_host_list) == 0:
                target_host_list = list(target_parameter.keys())
                target_host_list.sort()
            else:
                target_host_list.extend(list(target_parameter.keys()))
                target_host_list = list(set(target_host_list))
        target_host_list.sort()
        for host_name in target_host_list:
            tmp_input_order = list(target_parameter.get(host_name).keys())
            tmp_input_order = list(set(tmp_input_order))
            input_order_list["menu_1"].setdefault(host_name, tmp_input_order)
            input_order_list["menu_2"].setdefault(host_name, tmp_input_order)

        # set target_host_list input_order_list
        compare_config["target_host_list"] = target_host_list
        compare_config["input_order_list"] = input_order_list

        # complement target_parameter
        compare_config = _complement_target_parameter(compare_config, target_parameter)

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "compare_target_menu_error", True)
        retBool = False

    return retBool, compare_config,


# 補完用対象項目設定(rest_name)
def _set_base_parameters(compare_config, restkey_list, target_key):
    """
        set base parameters [caller:_get_target_datas]
        ARGS:
            compare_config, restkey_list, target_key
        RETRUN:
            {}
    """

    del_parameter_list = compare_config.get("del_parameter_list")

    # all rest_key
    base_parameter_all = {}
    for rest_key in restkey_list:
        base_parameter_all.setdefault(rest_key, None)
    compare_config["base_parameter_all"].setdefault(target_key, base_parameter_all)

    # id + parameter  rest_key
    base_parameter = {}
    for del_key in del_parameter_list:
        restkey_list.remove(del_key)
    for rest_key in restkey_list:
        base_parameter.setdefault(rest_key, None)
        compare_config["base_parameter"].setdefault(target_key, base_parameter)

    return compare_config


# 対象データ非対称補完
def _complement_target_parameter(compare_config, target_parameter):
    """
        complement target parameter [caller:_get_target_datas]
        ARGS:
            compare_config, restkey_list, target_key
        RETRUN:
            {}
    """
    target_host_list = compare_config.get("target_host_list")
    input_order_list = compare_config.get("input_order_list")
    base_parameter_all_1 = compare_config["base_parameter_all"]["menu_1"]
    base_parameter_all_2 = compare_config["base_parameter_all"]["menu_2"]

    for host_name in target_host_list:
        if host_name in target_parameter:
            tmp_io_list_1 = input_order_list["menu_1"][host_name]
            tmp_io_list_2 = input_order_list["menu_2"][host_name]
            # complement base menu1 input_order
            for tmp_io in tmp_io_list_1:
                if tmp_io in target_parameter[host_name]:
                    target_parameter[host_name].setdefault(tmp_io, {})
                    if "menu_1" not in target_parameter[host_name][tmp_io]:
                        target_parameter[host_name][tmp_io].setdefault("menu_1", base_parameter_all_1)
                    elif target_parameter[host_name][tmp_io]["menu_1"] is None:
                        target_parameter[host_name][tmp_io].setdefault("menu_1", base_parameter_all_1)
                    if "menu_2" not in target_parameter[host_name][tmp_io]:
                        target_parameter[host_name][tmp_io].setdefault("menu_2", base_parameter_all_2)
                    elif target_parameter[host_name][tmp_io]["menu_2"] is None:
                        target_parameter[host_name][tmp_io].setdefault("menu_2", base_parameter_all_2)
            # complement base menu2 input_order
            for tmp_io in tmp_io_list_2:
                if tmp_io not in target_parameter[host_name]:
                    target_parameter[host_name].setdefault(tmp_io, {})
                    if "menu_1" not in target_parameter[host_name][tmp_io]:
                        target_parameter[host_name][tmp_io].setdefault("menu_1", base_parameter_all_1)
                    elif target_parameter[host_name][tmp_io]["menu_1"] is None:
                        target_parameter[host_name][tmp_io].setdefault("menu_1", base_parameter_all_1)
                    if "menu_2" not in target_parameter[host_name][tmp_io]:
                        target_parameter[host_name][tmp_io].setdefault("menu_2", base_parameter_all_2)
                    elif target_parameter[host_name][tmp_io]["menu_2"] is None:
                        target_parameter[host_name][tmp_io].setdefault("menu_2", base_parameter_all_2)

    compare_config["origin_data"] = target_parameter

    return compare_config


# 対象データ整形
def _set_target_parameter(compare_config, target_key, vertical, filter_result):
    """
        set target parameter [caller:_get_target_datas]
        ARGS:
            compare_config, restkey_list, target_key
        RETRUN:
            {}
    """
    target_parameter = compare_config.get("_target_parameter")
    target_menu_info = compare_config.get("target_menu_info")

    if len(filter_result) >= 1:
        for row in filter_result:
            override_flg = False
            insert_flg = False
            tmp_parameter = row.get("parameter")
            host_name = tmp_parameter.get("host_name")
            target_menu_info.get(target_key).get("vertical")
            base_datetime = tmp_parameter.get("base_datetime")

            # vertical off
            input_order = "__no_input_order__"
            # vertical on
            if vertical == "1":
                input_order = tmp_parameter.get("input_order")

            # set insert_flg
            # set override_flg: input_order重複時, base_datetimeが最新のものを優先してSET
            target_parameter.setdefault(host_name, {})
            if input_order in target_parameter[host_name]:
                if target_key in target_parameter[host_name][input_order]:
                    p_base_datetime = target_parameter[host_name][input_order][target_key].get("base_datetime")
                    if base_datetime > p_base_datetime:
                        override_flg = True
                else:
                    insert_flg = True
            else:
                insert_flg = True

            # data set or override
            if insert_flg is True:
                target_parameter.setdefault(host_name, {})
                target_parameter[host_name].setdefault(input_order, {})
                target_parameter[host_name][input_order].setdefault(target_key, tmp_parameter)
            if override_flg is True:
                target_parameter[host_name][input_order][target_key] = tmp_parameter

        compare_config["_target_parameter"] = target_parameter

    return compare_config


# 比較設定取得
def _set_compare_config_info(objdbca, compare_config, options):
    """
        set compare_config_info  [caller:compare_execute]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    retBool = True

    try:
        # set compare_config
        retBool, compare_config = _set_compare_config(objdbca, compare_config, options)
        detail_flg = _get_flg(compare_config, "detail_flg")

        # set compare_detail_config
        if detail_flg is True:
            retBool, compare_config = _set_compare_detail_config(objdbca, compare_config, options)

        # set column_info
        retBool, compare_config = _set_column_info(objdbca, compare_config, options)

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "compare_target_menu_error", True)
        retBool = False

    return retBool, compare_config,


# 比較設定取得
def _set_compare_config(objdbca, compare_config, options):
    """
        set compare_config [caller:_set_compare_config_info]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    retBool = True

    try:
        # get language
        language = compare_config.get("language")
        # get compare_list
        compare_name = compare_config.get("parameter").get("compare")
        sql_str = textwrap.dedent("""
            SELECT
                `TAB_A`.*,
                `TAB_B`.`MENU_NAME_JA` AS `TARGET_MENU_NAME_1_JA`,
                `TAB_B`.`MENU_NAME_EN` AS `TARGET_MENU_NAME_1_EN`,
                `TAB_B`.`MENU_NAME_REST` AS `TARGET_MENU_NAME_1_REST`,
                `TAB_D`.`VERTICAL` AS `TARGET_MENU_1_VERTICAL`,
                `TAB_C`.`MENU_NAME_JA` AS `TARGET_MENU_NAME_2_JA`,
                `TAB_C`.`MENU_NAME_EN` AS `TARGET_MENU_NAME_2_EN`,
                `TAB_C`.`MENU_NAME_REST` AS `TARGET_MENU_NAME_2_REST`,
                `TAB_E`.`VERTICAL` AS `TARGET_MENU_2_VERTICAL`
            FROM
                `T_COMPARE_CONFG_LIST` `TAB_A`
            LEFT JOIN `T_COMN_MENU` `TAB_B` ON ( `TAB_A`.`TARGET_MENU_1` = `TAB_B`.`MENU_ID` )
            LEFT JOIN `T_COMN_MENU` `TAB_C` ON ( `TAB_A`.`TARGET_MENU_2` = `TAB_C`.`MENU_ID` )
            LEFT JOIN `T_COMN_MENU_TABLE_LINK` `TAB_D` ON ( `TAB_A`.`TARGET_MENU_1` = `TAB_D`.`MENU_ID` )
            LEFT JOIN `T_COMN_MENU_TABLE_LINK` `TAB_E` ON ( `TAB_A`.`TARGET_MENU_2` = `TAB_E`.`MENU_ID` )
            WHERE `TAB_A`.`COMPARE_NAME` = %s
            AND `TAB_A`.`DISUSE_FLAG` <> 1
            AND `TAB_B`.`DISUSE_FLAG` <> 1
            AND `TAB_C`.`DISUSE_FLAG` <> 1
            AND `TAB_D`.`DISUSE_FLAG` <> 1
            AND `TAB_E`.`DISUSE_FLAG` <> 1
        """).format().strip()
        bind_list = [compare_name]

        rows = objdbca.sql_execute(sql_str, bind_list)
        if len(rows) == 1:
            config_compare_list = rows[0]

            detail_flg = bool(strtobool(config_compare_list.get("DETAIL_FLAG")))
            compare_config = _set_flg(compare_config, "detail_flg", detail_flg)
            compare_config["config_compare_list"] = config_compare_list
            compare_config["target_menus"]["menu_1"] = config_compare_list.get("TARGET_MENU_NAME_1_REST")
            compare_config["target_menus"]["menu_2"] = config_compare_list.get("TARGET_MENU_NAME_2_REST")
            compare_config["target_menus_lang"]["menu_1"] = config_compare_list.get("TARGET_MENU_NAME_1_" + language.upper())
            compare_config["target_menus_lang"]["menu_2"] = config_compare_list.get("TARGET_MENU_NAME_2_" + language.upper())

            target_menu_info = {}
            target_menu_info.setdefault("menu_1", {})
            target_menu_info.setdefault("menu_2", {})
            target_menu_info["menu_1"].setdefault("menu_id", config_compare_list.get("TARGET_MENU_1"))
            target_menu_info["menu_2"].setdefault("menu_id", config_compare_list.get("TARGET_MENU_2"))
            target_menu_info["menu_1"].setdefault("menu_name_ja", config_compare_list.get("TARGET_MENU_NAME_1_JA"))
            target_menu_info["menu_2"].setdefault("menu_name_ja", config_compare_list.get("TARGET_MENU_NAME_2_JA"))
            target_menu_info["menu_1"].setdefault("menu_name_en", config_compare_list.get("TARGET_MENU_NAME_1_EN"))
            target_menu_info["menu_2"].setdefault("menu_name_en", config_compare_list.get("TARGET_MENU_NAME_2_EN"))
            target_menu_info["menu_1"].setdefault("menu_name_rest", config_compare_list.get("TARGET_MENU_NAME_1_REST"))
            target_menu_info["menu_2"].setdefault("menu_name_rest", config_compare_list.get("TARGET_MENU_NAME_2_REST"))
            target_menu_info["menu_1"].setdefault("vertical", config_compare_list.get("TARGET_MENU_1_VERTICAL"))
            target_menu_info["menu_2"].setdefault("vertical", config_compare_list.get("TARGET_MENU_2_VERTICAL"))
            compare_config["target_menu_info"] = target_menu_info

            vertical_1 = target_menu_info.get("menu_1").get("vertical")
            vertical_2 = target_menu_info.get("menu_2").get("vertical")
            # set vertical_compare_flg
            if vertical_1 == "1" and vertical_2 == "1":
                compare_config = _set_flg(compare_config, "vertical_compare_flg", True)
            elif vertical_1 == "1" and vertical_1 != vertical_2:
                # compare_config faild nomal * vertical
                target_key = "compare_config"
                msg_code = "MSG-60001"
                err_msg = g.appmsg.get_api_message(msg_code, [])
                compare_config = _set_msg(compare_config, target_key, err_msg)
                status_code = "499-01002"
                log_msg_args = [compare_name]
                api_msg_args = [compare_name]
                raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

        else:
            target_key = "compare_config"
            err_msg = "compare_config faild"
            compare_config = _set_msg(compare_config, target_key, err_msg)
            status_code = "499-01002"
            log_msg_args = [compare_name]
            api_msg_args = [compare_name]
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405
    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "compare_config_error", True)
        retBool = False

    return retBool, compare_config,


# 比較設定詳細取得
def _set_compare_detail_config(objdbca, compare_config, options):
    """
        set compare_detail_config  [caller:_set_compare_config_info]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    retBool = True

    try:
        # get compare_detail
        compare_name = compare_config.get("parameter").get("compare")
        sql_str = textwrap.dedent("""
            SELECT
                `TAB_A`.*,
                CONCAT(
                    `TBL_G`.`FULL_COL_GROUP_NAME_JA`,
                    '/',
                    `TAB_C`.`COLUMN_NAME_JA`
                ) AS `TARGET_COLUMN_NAME_1_JA`,
                CONCAT(
                    `TBL_H`.`FULL_COL_GROUP_NAME_EN`,
                    '/',
                    `TAB_C`.`COLUMN_NAME_EN`
                ) AS `TARGET_COLUMN_NAME_1_EN`,
                `TAB_C`.`COLUMN_NAME_REST` AS `TARGET_COLUMN_NAME_1_REST`,
                `TAB_C`.`COLUMN_CLASS` AS `TARGET_COLUMN_CLASS_1`,
                `TAB_E`.`COLUMN_CLASS_NAME` AS `TARGET_COLUMN_CLASS_NAME_1`,
                CONCAT(
                    `TBL_G`.`FULL_COL_GROUP_NAME_JA`,
                    '/',
                    `TAB_C`.`COLUMN_NAME_JA`
                ) AS `TARGET_COLUMN_NAME_2_JA`,
                CONCAT(
                    `TBL_H`.`FULL_COL_GROUP_NAME_EN`,
                    '/',
                    `TAB_C`.`COLUMN_NAME_EN`
                ) AS `TARGET_COLUMN_NAME_2_EN`,
                `TAB_D`.`COLUMN_NAME_REST` AS `TARGET_COLUMN_NAME_2_REST`,
                `TAB_D`.`COLUMN_CLASS` AS `TARGET_COLUMN_CLASS_2`,
                `TAB_F`.`COLUMN_CLASS_NAME` AS `TARGET_COLUMN_CLASS_NAME_2`
            FROM
                `T_COMPARE_DETAIL_LIST` `TAB_A`
            LEFT JOIN `T_COMPARE_CONFG_LIST` `TAB_B` ON ( `TAB_A`.`COMPARE_ID` = `TAB_B`.`COMPARE_ID` )
            LEFT JOIN `T_COMN_MENU_COLUMN_LINK` `TAB_C` ON ( `TAB_A`.`TARGET_COLUMN_ID_1` = `TAB_C`.`COLUMN_DEFINITION_ID` )
            LEFT JOIN `T_COMN_MENU_COLUMN_LINK` `TAB_D` ON ( `TAB_A`.`TARGET_COLUMN_ID_2` = `TAB_D`.`COLUMN_DEFINITION_ID` )
            LEFT JOIN `T_COMN_COLUMN_CLASS` `TAB_E` ON ( `TAB_E`.`COLUMN_CLASS_ID` = `TAB_C`.`COLUMN_CLASS` )
            LEFT JOIN `T_COMN_COLUMN_CLASS` `TAB_F` ON ( `TAB_F`.`COLUMN_CLASS_ID` = `TAB_D`.`COLUMN_CLASS` )
            LEFT JOIN `T_COMN_COLUMN_GROUP` `TBL_G` ON ( `TAB_C`.`COL_GROUP_ID` = `TBL_G`.`COL_GROUP_ID` )
            LEFT JOIN `T_COMN_COLUMN_GROUP` `TBL_H` ON ( `TAB_D`.`COL_GROUP_ID` = `TBL_H`.`COL_GROUP_ID` )
            WHERE `TAB_B`.`COMPARE_NAME` = %s
            AND `TAB_A`.`DISUSE_FLAG` <> 1
            AND `TAB_B`.`DISUSE_FLAG` <> 1
            AND `TAB_C`.`DISUSE_FLAG` <> 1
            AND `TAB_D`.`DISUSE_FLAG` <> 1
            AND `TBL_G`.`DISUSE_FLAG` <> 1
            AND `TBL_H`.`DISUSE_FLAG` <> 1
            ORDER BY `TAB_A`.`DISP_SEQ` ASC
        """).format().strip()
        bind_list = [compare_name]

        rows = objdbca.sql_execute(sql_str, bind_list)
        config_compare_detail = {}
        config_compare_disp = []

        if len(rows) >= 1:
            row = {
                'COMPARE_DETAIL_ID': 'menu',
                'COMPARE_ID': 'menu',
                'COMPARE_COL_TITLE': g.appmsg.get_api_message('MSG-60028'),  # 'メニュー名',
                'TARGET_COLUMN_ID_1': 'menu',
                'TARGET_INPUT_ORDER_1': None,
                'TARGET_COLUMN_ID_2': 'menu',
                'TARGET_INPUT_ORDER_2': None,
                'DISP_SEQ': 0,
                'NOTE': None,
                'DISUSE_FLAG': '0',
                'TARGET_COLUMN_NAME_1_JA': 'メニュー名',
                'TARGET_COLUMN_NAME_1_EN': 'menu',
                'TARGET_COLUMN_NAME_1_REST': 'menu',
                'TARGET_COLUMN_CLASS_1': '1',
                'TARGET_COLUMN_CLASS_NAME_1': 'SingleTextColumn',
                'TARGET_COLUMN_NAME_2_JA': 'メニュー名',
                'TARGET_COLUMN_NAME_2_EN': 'menu',
                'TARGET_COLUMN_NAME_2_REST': 'menu',
                'TARGET_COLUMN_CLASS_2': '1',
                'TARGET_COLUMN_CLASS_NAME_2': 'SingleTextColumn'
            }
            compare_col_title = row.get("COMPARE_COL_TITLE")
            config_compare_detail.setdefault(compare_col_title, row)
            config_compare_disp.append(row)
            row = {
                'COMPARE_DETAIL_ID': 'uuid',
                'COMPARE_ID': 'uuid',
                'COMPARE_COL_TITLE': g.appmsg.get_api_message('MSG-60029'),  # '項番',
                'TARGET_COLUMN_ID_1': 'uuid',
                'TARGET_INPUT_ORDER_1': None,
                'TARGET_COLUMN_ID_2': 'uuid',
                'TARGET_INPUT_ORDER_2': None,
                'DISP_SEQ': 0,
                'NOTE': None,
                'DISUSE_FLAG': '0',
                'TARGET_COLUMN_NAME_1_JA': '項番',
                'TARGET_COLUMN_NAME_1_EN': 'uuid',
                'TARGET_COLUMN_NAME_1_REST': 'uuid',
                'TARGET_COLUMN_CLASS_1': '1',
                'TARGET_COLUMN_CLASS_NAME_1': 'SingleTextColumn',
                'TARGET_COLUMN_NAME_2_JA': '項番',
                'TARGET_COLUMN_NAME_2_EN': 'uuid',
                'TARGET_COLUMN_NAME_2_REST': 'uuid',
                'TARGET_COLUMN_CLASS_2': '1',
                'TARGET_COLUMN_CLASS_NAME_2': 'SingleTextColumn'
            }
            compare_col_title = row.get("COMPARE_COL_TITLE")
            config_compare_detail.setdefault(compare_col_title, row)
            config_compare_disp.append(row)
            row = {
                'COMPARE_DETAIL_ID': 'operation_name_disp',
                'COMPARE_ID': 'operation_name_disp',
                'COMPARE_COL_TITLE': f"{g.appmsg.get_api_message('MSG-90046')}/{g.appmsg.get_api_message('MSG-60030')}",  # 'オペレーション/オペレーション名',
                'TARGET_COLUMN_ID_1': 'operation_name_disp',
                'TARGET_INPUT_ORDER_1': None,
                'TARGET_COLUMN_ID_2': 'operation_name_disp',
                'TARGET_INPUT_ORDER_2': None,
                'DISP_SEQ': 0,
                'NOTE': None,
                'DISUSE_FLAG': '0',
                'TARGET_COLUMN_NAME_1_JA': 'オペレーション名',
                'TARGET_COLUMN_NAME_1_EN': 'operation_name_disp',
                'TARGET_COLUMN_NAME_1_REST': 'operation_name_disp',
                'TARGET_COLUMN_CLASS_1': '1',
                'TARGET_COLUMN_CLASS_NAME_1': 'SingleTextColumn',
                'TARGET_COLUMN_NAME_2_JA': 'オペレーション名',
                'TARGET_COLUMN_NAME_2_EN': 'operation_name_disp',
                'TARGET_COLUMN_NAME_2_REST': 'operation_name_disp',
                'TARGET_COLUMN_CLASS_2': '1',
                'TARGET_COLUMN_CLASS_NAME_2': 'SingleTextColumn'
            }
            compare_col_title = row.get("COMPARE_COL_TITLE")
            config_compare_detail.setdefault(compare_col_title, row)
            config_compare_disp.append(row)
            row = {
                'COMPARE_DETAIL_ID': 'base_datetime',
                'COMPARE_ID': 'base_datetime',
                'COMPARE_COL_TITLE': f"{g.appmsg.get_api_message('MSG-90046')}/{g.appmsg.get_api_message('MSG-60031')}",  # 'オペレーション/基準日時',
                'TARGET_COLUMN_ID_1': 'base_datetime',
                'TARGET_INPUT_ORDER_1': None,
                'TARGET_COLUMN_ID_2': 'base_datetime',
                'TARGET_INPUT_ORDER_2': None,
                'DISP_SEQ': 0,
                'NOTE': None,
                'DISUSE_FLAG': '0',
                'TARGET_COLUMN_NAME_1_JA': '基準日時',
                'TARGET_COLUMN_NAME_1_EN': 'base_datetime',
                'TARGET_COLUMN_NAME_1_REST': 'base_datetime',
                'TARGET_COLUMN_CLASS_1': '1',
                'TARGET_COLUMN_CLASS_NAME_1': 'SingleTextColumn',
                'TARGET_COLUMN_NAME_2_JA': '基準日時',
                'TARGET_COLUMN_NAME_2_EN': 'base_datetime',
                'TARGET_COLUMN_NAME_2_REST': 'base_datetime',
                'TARGET_COLUMN_CLASS_2': '1',
                'TARGET_COLUMN_CLASS_NAME_2': 'SingleTextColumn'
            }
            compare_col_title = row.get("COMPARE_COL_TITLE")
            config_compare_detail.setdefault(compare_col_title, row)
            config_compare_disp.append(row)
            for row in rows:
                # disq_seq = row.get("DISP_SEQ")
                compare_col_title = row.get("COMPARE_COL_TITLE")
                config_compare_detail.setdefault(compare_col_title, row)
                config_compare_disp.append(row)
            compare_config["config_compare_detail"] = config_compare_detail
            compare_config["config_compare_disp"] = config_compare_disp
        else:
            target_key = "compare_config"
            err_msg = "compare_config_detail faild"
            compare_config = _set_msg(compare_config, target_key, err_msg)
            status_code = "499-01003"
            log_msg_args = []
            api_msg_args = []
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "compare_config_error", True)
        retBool = False

    return retBool, compare_config,


# 比較項目設定
def _set_column_info(objdbca, compare_config, options):
    """
        set column_info  [caller:_set_compare_config_info]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    retBool = True
    language = compare_config.get("language")

    try:
        column_info = []
        file_column_list = [9, "9", "FileUploadColumn", "FileUploadEncryptColumn"]
        detail_flg = _get_flg(compare_config, "detail_flg")
        target_menu_info = compare_config.get("target_menu_info")
        del_parameter_list = compare_config.get("del_parameter_list")
        no_compare_parameter_list = compare_config.get("no_compare_parameter_list")
        no_input_order_list = compare_config.get("no_input_order_list")
        vertical_no_use_rest_key = compare_config.get("vertical_no_use_rest_key")

        compare_config = _set_flg(compare_config, "compare_file", False)

        # detail_flg: True: use T_COMPARE_DETAIL_LIST / False: use T_COMN_MENU_COLUMN_LINK
        if detail_flg is False:
            menu_id = target_menu_info.get("menu_1").get("menu_id")
            menu_id_2 = target_menu_info.get("menu_2").get("menu_id")

            sql_str = textwrap.dedent("""
                SELECT
                    `TAB_A`.*,
                    `TAB_B`.`COLUMN_CLASS_NAME`,
                    `TAB_C`.*
                FROM
                    `T_COMN_MENU_COLUMN_LINK` `TAB_A`
                LEFT JOIN `T_COMN_COLUMN_CLASS` `TAB_B` ON ( `TAB_A`.`COLUMN_CLASS` = `TAB_B`.`COLUMN_CLASS_ID` )
                LEFT JOIN `T_COMN_COLUMN_GROUP` `TAB_C` ON ( `TAB_A`.`COL_GROUP_ID` = `TAB_C`.`COL_GROUP_ID` )
                WHERE `TAB_A`.`MENU_ID` = %s
                AND `TAB_A`.`DISUSE_FLAG` <> 1
                AND `TAB_C`.`DISUSE_FLAG` <> 1
                ORDER BY  `TAB_A`.`COLUMN_DISP_SEQ` ASC
            """).format().strip()
            bind_list = [menu_id]
            # compare_list
            rows = objdbca.sql_execute(sql_str, bind_list)

            sql_str_2 = textwrap.dedent("""
                SELECT
                    `TAB_A`.*,
                    `TAB_B`.`COLUMN_CLASS_NAME`
                FROM
                    `T_COMN_MENU_COLUMN_LINK` `TAB_A`
                LEFT JOIN `T_COMN_COLUMN_CLASS` `TAB_B` ON ( `TAB_A`.`COLUMN_CLASS` = `TAB_B`.`COLUMN_CLASS_ID` )
                WHERE `TAB_A`.`MENU_ID` = %s
                AND `TAB_A`.`DISUSE_FLAG` <> 1
                ORDER BY  `TAB_A`.`COLUMN_DISP_SEQ` ASC
            """).format().strip()
            bind_list_2 = [menu_id_2]
            # compare_list
            rows_2 = objdbca.sql_execute(sql_str_2, bind_list_2)

            if len(rows) >= 1:
                row = {
                    'COMPARE_DETAIL_ID': 'menu',
                    'COMPARE_ID': 'menu',
                    'COMPARE_COL_TITLE': 'メニュー名',
                    'TARGET_COLUMN_ID_1': 'menu',
                    'TARGET_INPUT_ORDER_1': None,
                    'TARGET_COLUMN_ID_2': 'menu',
                    'TARGET_INPUT_ORDER_2': None,
                    'DISP_SEQ': 0,
                    'NOTE': None,
                    'DISUSE_FLAG': '0',
                    'TARGET_COLUMN_NAME_1_JA': 'メニュー名',
                    'TARGET_COLUMN_NAME_1_EN': 'menu',
                    'TARGET_COLUMN_NAME_1_REST': 'menu',
                    'TARGET_COLUMN_CLASS_1': '1',
                    'TARGET_COLUMN_CLASS_NAME_1': 'SingleTextColumn',
                    'TARGET_COLUMN_NAME_2_JA': 'メニュー名',
                    'TARGET_COLUMN_NAME_2_EN': 'menu',
                    'TARGET_COLUMN_NAME_2_REST': 'menu',
                    'TARGET_COLUMN_CLASS_2': '1',
                    'TARGET_COLUMN_CLASS_NAME_2': 'SingleTextColumn',
                    'COLUMN_NAME_JA': 'メニュー名',
                    'COLUMN_NAME_EN': 'menu',
                    'COLUMN_NAME_REST': 'menu'
                }
                # tmp_rows = []
                # tmp_rows.append(row)
                # for row in rows:
                #    tmp_rows.append(row)
                # for row in tmp_rows:

                for row in rows:
                    for row_2 in rows_2:
                        col_key_2 = row_2.get("COLUMN_NAME_REST")
                        if col_key_2 not in del_parameter_list:
                            if row_2.get("COLUMN_CLASS") == row.get("COLUMN_CLASS"):
                                column_class_2 = row_2.get("COLUMN_CLASS")
                                column_class_2 = row_2.get("COLUMN_CLASS_NAME")

                    compare_col_title = "{}/{}".format(
                        row.get("FULL_COL_GROUP_NAME_" + language.upper()),
                        row.get("COLUMN_NAME_" + language.upper())
                    )
                    col_key = row.get("COLUMN_NAME_REST")
                    if col_key not in del_parameter_list:
                        column_class = row.get("COLUMN_CLASS")
                        column_class = row.get("COLUMN_CLASS_NAME")
                        # file comapre target flg
                        file_flg = False
                        if column_class in file_column_list or column_class_2 in file_column_list:
                            compare_config = _set_flg(compare_config, "compare_file", True)
                            file_flg = True

                        # comapre target flg
                        compare_target_flg = False
                        if col_key not in no_compare_parameter_list:
                            compare_target_flg = True

                        # no input order flg
                        no_input_order_flg = False
                        if col_key in no_input_order_list:
                            no_input_order_flg = True

                        # vertical no use flg
                        vertical_no_use_flg = False
                        if col_key in vertical_no_use_rest_key:
                            vertical_no_use_flg = True

                        tmp_column_info = {
                            "col_name": compare_col_title,
                            "menu_1": col_key,
                            "menu_2": col_key,
                            "file_flg": file_flg,
                            "compare_target_flg": compare_target_flg,
                            "no_input_order_flg": no_input_order_flg,
                            "vertical_no_use_flg": vertical_no_use_flg,
                            "column_class_1": column_class,
                            "column_class_2": column_class_2
                        }
                        column_info.append(tmp_column_info)
        else:
            config_compare_disp = compare_config.get("config_compare_disp")
            if len(config_compare_disp) >= 1:
                for column_row in config_compare_disp:
                    compare_col_title = column_row.get("COMPARE_COL_TITLE")
                    col_key_1 = column_row.get("TARGET_COLUMN_NAME_1_REST")
                    col_key_2 = column_row.get("TARGET_COLUMN_NAME_2_REST")
                    column_class_1 = column_row.get("TARGET_COLUMN_CLASS_1")
                    column_class_2 = column_row.get("TARGET_COLUMN_CLASS_2")
                    column_class_1 = column_row.get("TARGET_COLUMN_CLASS_NAME_1")
                    column_class_2 = column_row.get("TARGET_COLUMN_CLASS_NAME_2")
                    input_order_1 = column_row.get("TARGET_INPUT_ORDER_1")
                    input_order_2 = column_row.get("TARGET_INPUT_ORDER_2")
                    if input_order_1 is None and input_order_2 is None:
                        input_order_1 = "__no_input_order__"
                        input_order_2 = "__no_input_order__"

                    # file comapre target flg
                    file_flg = False
                    if column_class_1 in file_column_list or column_class_2 in file_column_list:
                        compare_config = _set_flg(compare_config, "compare_file", True)
                        file_flg = True

                    # comapre target flg
                    compare_target_flg = True
                    no_input_order_flg = False
                    vertical_no_use_flg = False
                    if col_key_1 is None and col_key_2 is None:
                        compare_target_flg = False
                    else:
                        # comapre target flg
                        if col_key_1 in no_compare_parameter_list and col_key_2 in no_compare_parameter_list:
                            compare_target_flg = False
                        # no input order flg
                        if col_key_1 in no_input_order_list and col_key_2 in no_input_order_list:
                            no_input_order_flg = True
                        # vertical no use flg
                        if col_key_1 in vertical_no_use_rest_key and col_key_2 in vertical_no_use_rest_key:
                            vertical_no_use_flg = True

                    tmp_column_info = {
                        "col_name": compare_col_title,
                        "menu_1": col_key_1,
                        "menu_2": col_key_2,
                        "file_flg": file_flg,
                        "compare_target_flg": compare_target_flg,
                        "no_input_order_flg": no_input_order_flg,
                        "vertical_no_use_flg": vertical_no_use_flg,
                        "column_class_1": column_class_1,
                        "column_class_2": column_class_2,
                        "input_order_1": input_order_1,
                        "input_order_2": input_order_2,
                    }
                    column_info.append(tmp_column_info)

        # set column_info
        compare_config["column_info"] = column_info
    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "compare_target_menu_error", True)
        retBool = False

    return retBool, compare_config,


# 比較項目設定(縦メニュー変換 item_1 -> item_1[input_order:X]
def _override_column_info_vertical_ver(objdbca, compare_config, options):
    """
        override column_info vertical ver  [caller:compare_execute]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    retBool = True
    base_column_info = compare_config.get("column_info")
    input_order_lang = compare_config.get("input_order_lang")

    try:
        column_info = []
        origin_data = compare_config.get("origin_data")
        vertical_compare_flg = _get_flg(compare_config, "vertical_compare_flg")

        if vertical_compare_flg is True:
            input_orders = []
            for target_host in origin_data:
                tmp_input_orders = list(origin_data[target_host].keys())
                tmp_input_orders = list(set(tmp_input_orders))
                input_orders.extend(tmp_input_orders)
                input_orders = list(set(input_orders))

            if "__no_input_order__" in input_orders:
                input_orders.remove("__no_input_order__")
            input_orders.sort()

            # column_info.append(base_column)

            for base_column in base_column_info:
                col_name = base_column["col_name"]
                no_input_order_flg = base_column["no_input_order_flg"]
                vertical_no_use_flg = base_column["vertical_no_use_flg"]
                if no_input_order_flg is False:
                    for input_order in input_orders:
                        tmp_base_column = base_column.copy()
                        tmp_base_column["col_name"] = "{}[{}:{}]".format(col_name, input_order_lang, input_order)
                        tmp_base_column["col_name"] = _get_col_name_input_order(col_name, input_order, input_order_lang)
                        # set column_info
                        column_info.append(tmp_base_column)
                elif vertical_no_use_flg is True:
                    # if col_name in ["menu", "メニュー名"]:
                    #     column_info.append(base_column)
                    pass
                else:
                    # set column_info
                    column_info.append(base_column)
            compare_config["column_info"] = column_info

        else:
            pass

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "compare_target_menu_error", True)
        retBool = False

    return retBool, compare_config,


# パラメータチェック
def _chk_parameters(objdbca, compare_config, options):
    """
        check parameters: compare,base_date,other_options  [caller:compare_execute]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    retBool = True

    try:
        # _chk_parameter_compare
        compare_config = _chk_parameter_compare(objdbca, compare_config, options)
        # _chk_parameter_base_date
        compare_config = _chk_parameter_base_date(objdbca, compare_config, options)
        # _chk_parameter_host
        compare_config = _chk_parameter_host(objdbca, compare_config, options)
        # _chk_parameter_copmare_target_column
        compare_config = _chk_parameter_copmare_target_column(objdbca, compare_config, options)
        # chk_compare
        compare_config = _chk_parameter_other_options(objdbca, compare_config, options)

        # error flg check
        tmp_err_flg = _get_flg(compare_config, "parameter_error")
        if tmp_err_flg is True:
            err_msg = _get_msg(compare_config, "parameter_error")
            status_code = "499-01000"
            log_msg_args = [err_msg]
            api_msg_args = [err_msg]
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        retBool = False

    return retBool, compare_config,


#  パラメータチェック:compare
def _chk_parameter_compare(objdbca, compare_config, options):
    """
        check parameter:compare [caller:_chk_parameters]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    try:
        tmp_parameter = compare_config.get("parameter")
        compare = tmp_parameter.get("compare")

        if compare:
            sql_str = textwrap.dedent("""
                SELECT * FROM `T_COMPARE_CONFG_LIST` `TAB_A`
                WHERE `TAB_A`.`COMPARE_NAME` = %s
                AND `TAB_A`.`DISUSE_FLAG` <> 1
            """).format().strip()
            bind_list = [compare]
            rows = objdbca.sql_execute(sql_str, bind_list)
            if len(rows) != 1:
                target_key = "compare"
                # compare faild
                msg_code = "MSG-60005"
                msg_args = [target_key, compare]
                err_msg = g.appmsg.get_api_message(msg_code, msg_args)
                compare_config = _set_msg(compare_config, target_key, err_msg)

                status_code = "499-01001"
                log_msg_args = [err_msg]
                api_msg_args = [err_msg]
                raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405
        else:
            target_key = "compare"
            # compare faild
            msg_code = "MSG-60005"
            msg_args = [target_key, compare]
            err_msg = g.appmsg.get_api_message(msg_code, msg_args)
            compare_config = _set_msg(compare_config, target_key, err_msg)
            status_code = "499-01001"
            log_msg_args = [err_msg]
            api_msg_args = [err_msg]
            raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "parameter_error", True)

    return compare_config


# パラメータチェック:base_date
def _chk_parameter_base_date(objdbca, compare_config, options):
    """
        check parameter:base_date [caller:_chk_parameters]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    try:
        tmp_parameter = compare_config.get("parameter")
        base_date_1 = tmp_parameter.get("base_date_1")
        base_date_2 = tmp_parameter.get("base_date_2")
        if base_date_1:
            base_date_1 = _chk_date_time_format("base_date_1", base_date_1)
            compare_config["parameter"]["base_date_1"] = base_date_1
        else:
            pass

        if base_date_2:
            base_date_2 = _chk_date_time_format("base_date_2", base_date_2)
            compare_config["parameter"]["base_date_2"] = base_date_2
        else:
            pass

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "parameter_error", True)

    return compare_config


# 日付形式のチェック
def _chk_date_time_format(key_name, val):
    """
        check datetime format [caller:_chk_parameters]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    # YYYY/MM/DD hh:mm:ssの場合OK
    if re.match(r'^[0-9]{4}/[0-9]{2}/[0-9]{2} [0-9]{2}:[0-9]{2}:[0-9]{2}$', val) is not None:
        pass
    # YYYY/MM/DD hh:mmの場合OK、:ssを補完
    elif re.match(r'^[0-9]{4}/[0-9]{2}/[0-9]{2} [0-9]{2}:[0-9]{2}$', val) is not None:
        val = val + ':00'
    # YYYY/MM/DDの場合OK、:hh:mm:ssを補完
    elif re.match(r'^[0-9]{4}/[0-9]{2}/[0-9]{2}$', val) is not None:
        val = val + ' 00:00:00'
    else:
        # date_time format error
        msg_code = "MSG-60006"
        msg_args = [key_name, val]
        err_msg = g.appmsg.get_api_message(msg_code, msg_args)
        status_code = "499-01001"
        log_msg_args = [err_msg]
        api_msg_args = [err_msg]
        raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405
    return val


# パラメータチェック:host
def _chk_parameter_host(objdbca, compare_config, options):
    """
        check parameter:host [caller:_chk_parameters]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    try:
        tmp_parameter = compare_config.get("parameter")
        host = tmp_parameter.get("host")

        search_hosts = []
        if isinstance(host, list):
            search_hosts = host
        elif isinstance(host, str):
            if host:
                search_hosts.append(host)
        else:
            if host is not None:
                target_key = "host"
                # host format error
                msg_code = "MSG-60006"
                msg_args = [target_key, host]
                err_msg = g.appmsg.get_api_message(msg_code, msg_args)
                compare_config = _set_msg(compare_config, target_key, err_msg)
                compare_config = _set_flg(compare_config, "parameter_error", True)
                status_code = "499-01001"
                log_msg_args = [err_msg]
                api_msg_args = [err_msg]
                raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

        if len(search_hosts) >= 1:
            for tmp_host in search_hosts:
                sql_str = textwrap.dedent("""
                    SELECT * FROM `T_ANSC_DEVICE` `TAB_A`
                    WHERE `TAB_A`.`HOST_NAME` = %s
                    AND `TAB_A`.`DISUSE_FLAG` <> 1
                """).format().strip()
                bind_list = [tmp_host]
                rows = objdbca.sql_execute(sql_str, bind_list)
                if len(rows) != 1:
                    target_key = "host"
                    # host value error
                    msg_code = "MSG-60005"
                    msg_args = [target_key, tmp_host]
                    err_msg = g.appmsg.get_api_message(msg_code, msg_args)
                    compare_config = _set_msg(compare_config, target_key, err_msg)
                    status_code = "499-01001"
                    log_msg_args = [err_msg]
                    api_msg_args = [err_msg]
                    raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "parameter_error", True)

    return compare_config


# パラメータチェック:copmare_target_column
def _chk_parameter_copmare_target_column(objdbca, compare_config, options):
    """
        check parameter:copmare_target_column [caller:_chk_parameters]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    try:
        tmp_parameter = compare_config.get("parameter")
        copmare_target_column = tmp_parameter.get("copmare_target_column")

        if copmare_target_column is None:
            pass
        elif isinstance(copmare_target_column, str):
            pass
        else:
            if copmare_target_column is not None:
                target_key = "copmare_target_column"
                # copmare_target_column value error
                msg_code = "MSG-60005"
                msg_args = [target_key, copmare_target_column]
                err_msg = g.appmsg.get_api_message(msg_code, msg_args)
                compare_config = _set_msg(compare_config, target_key, err_msg)
                compare_config = _set_flg(compare_config, "parameter_error", True)
                status_code = "499-01001"
                log_msg_args = [err_msg]
                api_msg_args = [err_msg]
                raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "parameter_error", True)

    return compare_config


# パラメータチェック:other_options
def _chk_parameter_other_options(objdbca, compare_config, options):
    """
        check parameter:other_options [caller:_chk_parameters]
        ARGS:
            compare_config, compare_config, options
        RETRUN:
            {}
    """
    try:
        tmp_parameter = compare_config.get("parameter")
        other_options = tmp_parameter.get("other_options")
        if isinstance(other_options, dict):
            # output_language
            if "output_language" in other_options:
                output_language = other_options.get("output_language")
                if output_language:
                    if output_language in ["en", "ja"]:
                        pass
                    else:
                        target_key = "other_options.output_language"
                        # language value error
                        msg_code = "MSG-60005"
                        msg_args = [target_key, output_language]
                        err_msg = g.appmsg.get_api_message(msg_code, msg_args)

                        status_code = "499-01001"
                        log_msg_args = [err_msg]
                        api_msg_args = [err_msg]
                        raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405
                else:
                    pass
            pass
            # check other options
        else:
            if other_options is not None:
                target_key = "other_options"
                # other_options format error
                msg_code = "MSG-60006"
                msg_args = [target_key, other_options]
                err_msg = g.appmsg.get_api_message(msg_code, msg_args)
                compare_config = _set_msg(compare_config, target_key, err_msg)
                compare_config = _set_flg(compare_config, "parameter_error", True)
                status_code = "499-01001"
                log_msg_args = [err_msg]
                api_msg_args = [err_msg]
                raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        compare_config = _set_flg(compare_config, "parameter_error", True)

    return compare_config


# ファイルデータ、mimetype取得
def _get_file_data_columnclass(objdbca, objtable, rest_key, file_name, target_uuid, col_class_name='TextColumn', file_required=False):
    """
        get file_data file_mimetype
        ARGS:
            objdbca, objtable, rest_key, file_name, target_uuid, col_class_name
        RETRUN:
            file_path(file_data), file_mimetype
    """
    file_data = ""
    file_mimetype = "text/plain"
    try:
        eval_class_str = "{}(objdbca,objtable,rest_key,'')".format(col_class_name)
        objcolumn = eval(eval_class_str)
        # ファイルの内容が必要な場合のみ、base64でファイルを取得
        if file_required is True or col_class_name != "FileUploadColumn":
            file_data = objcolumn.get_file_data(file_name, target_uuid, None)
        file_path = objcolumn.get_file_data_path(file_name, target_uuid, None)
        file_mimetype, encoding = mimetypes.guess_type(file_path, False)
        if file_mimetype is None:
            if os.path.exists(file_path):
                # check binary file
                ret, file_mimetype, encoding = no_mimetype_is_binary_chk(file_path, file_mimetype, encoding)
                if ret is False and file_mimetype is None:
                    file_mimetype = "text/plain"

    except AppException as _app_e:  # noqa: F405
        raise AppException(_app_e)  # noqa: F405
    except Exception:
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))

    if file_required is True or col_class_name != "FileUploadColumn":
        return file_data, file_mimetype
    return file_path, file_mimetype


# 代入順序付き項目名
def _get_col_name_input_order(col_name, input_order, input_order_lang=None):
    """
        get col_name add input_order
        ARGS:
            col_name, input_order, input_order_lang=None
        RETRUN:
            col_name_io
    """
    if input_order_lang is not None:
        col_name_io = "{}[{}:{}]".format(col_name, input_order_lang, input_order)
    elif isinstance(input_order, int):
        col_name_io = "{}[{}]".format(col_name, input_order)
    else:
        col_name_io = "{}".format(col_name)
    return col_name_io


# ファイル出力:Excel
def _create_outputfile(objdbca, compare_config, data, options):
    """
        create excel base64
        ARGS:
            objdbca, compare_config, data, options
        RETRUN:
            base64 string
    """

    try:
        # set
        tbl_start_str = "A"
        tbl_end_str = "E"
        work_dir_path = None

        # get data
        config = data.get("config")
        compare_data = data.get("compare_data")
        compare_diff_flg = data.get("compare_diff_flg")
        base_name = compare_config.get("parameter").get("compare")
        exec_time = datetime.datetime.now().strftime('%Y%m%d%H%M%S')  # noqa: F405
        tmp_exec_time = datetime.datetime.strptime(exec_time, '%Y%m%d%H%M%S')  # noqa: F405
        tmp_exec_time = tmp_exec_time.strftime('%Y/%m/%d %H:%M:%S')
        # exec_time = datetime.datetime.now().strftime('%Y%m%d')  # noqa: F405

        file_name = base_name + '_' + exec_time + '.xlsx'
        wbEncode = ""
        result = {
            "file_name": None,
            "file_data": None,
        }

        # get path
        work_dir_path, file_path, template_file_path = get_workdir_path(file_name)

        # create excel
        # set Workbook
        wb = Workbook()
        # wb = load_workbook(template_file_path)

        # set data: comapre execute parameters
        ws_index_create_table(wb, file_name, tmp_exec_time, config, compare_config, compare_diff_flg, tbl_start_str, tbl_end_str)

        # comapre result :target_host sheet
        ws_target_host_create_table(wb, file_name, tmp_exec_time, config, compare_data, compare_diff_flg, tbl_start_str, tbl_end_str)

        # create work dir
        if os.path.isdir(work_dir_path) is False:
            os.makedirs(work_dir_path, exist_ok=True)

        # save book
        wb.save(file_path)  # noqa: E303

    except AppException as _app_e:  # noqa: F405
        # clear work_dir
        if work_dir_path is not None and os.path.isdir(work_dir_path) is True:
            shutil.rmtree(work_dir_path)
        raise AppException(_app_e)  # noqa: F405
    except Exception as e:
        # clear work_dir
        if work_dir_path is not None and os.path.isdir(work_dir_path) is True:
            shutil.rmtree(work_dir_path)
        type_, value, traceback_ = sys.exc_info()
        msg = traceback.format_exception(type_, value, traceback_)
        g.applogger.info(addline_msg('{}{}'.format(msg, sys._getframe().f_code.co_name)))
        status_code = "499-01007"
        log_msg_args = [e]
        api_msg_args = [e]
        raise AppException(status_code, log_msg_args, api_msg_args)  # noqa: F405

    return file_path


# Excel入力値エスケープ処理
def wr_fmt(val):
    """
        excel writer format value escape
        ARGS:
            val
        RETRUN:
            escape(str(val))
    """
    if val is None:
        val = ""
    return escape(str(val))  # noqa: F405


# 文字列ランダム生成
def rand_str(n):
    """
        create random string length n :a-z
        ARGS:
            n
        RETRUN:
            string
    """
    rand_list = [chr(random.randint(97, 122)) for _ in range(n)]
    rand_chr = "".join(rand_list)
    return rand_chr


# 文字列長カウント
def count_string(val):
    """
        count value
        ARGS:
            val
        RETRUN:
            int
    """
    val_length = 0
    if isinstance(val, str):
        for tmp_val in val:
            if unicodedata.east_asian_width(tmp_val) in 'FWA':
                val_length += 2
            else:
                val_length += 1
    elif isinstance(val, int):
        val_length = len(str(val))

    return val_length


# 作業ディレクトリ取得
def get_workdir_path(file_name):
    """
        get work path
        ARGS:
            file_name
        RETRUN:
            work_dir_path, file_path, template_file_path,
    """

    tmp_work_dir = uuid.uuid4()

    # path: tmp work dir
    work_dir_path = "/tmp/{}".format(
        tmp_work_dir,
    ).replace('//', '/')

    # path: tmp work
    file_path = "/tmp/{}/{}".format(
        tmp_work_dir,
        file_name
    ).replace('//', '/')

    # path: tmplate xlsx
    template_file_name = "compare_base_template.xlsx"
    template_file_path = "/tmp/{}/{}".format(
        tmp_work_dir,
        template_file_name
    ).replace('//', '/')
    return work_dir_path, file_path, template_file_path,


# Excel:Cell幅調整
def ws_auto_adjusted_width(ws, default_length=8, limit_length=50, add_length=2, coefficient=1.2):
    """
        auto adjusted width
        ARGS:
            ws,  limit_length=50, default_length=8, add_length=2, coefficient=1.2
        RETRUN:
            ws
    """
    # set column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            tmp_length = count_string(cell.value)
            if tmp_length > max_length:
                max_length = tmp_length
        if default_length > max_length:
            max_length = default_length
        if max_length > limit_length:
            max_length = limit_length
        adjusted_width = (max_length + add_length) * coefficient
        column = get_column_letter(column)
        ws.column_dimensions[column].width = adjusted_width
    return ws


# Excel:テーブル作成
def ws_add_table(ws, tablename="", table_style=None, ref_area="A1:B1"):
    """
        create table
        ARGS:
            ws, tablename="", table_style=None, ref_area="A1:B1"
        RETRUN:
            ws
    """

    # default tablename
    if len(tablename) == 0:
        tablename = str("table") + rand_str(10)
    # default table area
    if len(ref_area) < 5:
        ref_area = "A1:B1"
    # default table stayle
    if table_style is None:
        table_style = TableStyleInfo(
            name='TableStyleMedium2',
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )

    table_data = Table(displayName=tablename, ref=ref_area)
    table_data.tableStyleInfo = table_style
    ws.add_table(table_data)

    return ws


# Excel:行入力
def ws_add_table_data_cells(ws, row_no, column_no, col_name="", val_1="", val_2="", str_diff="", other="", diff_flg=False, wrap_text_flg=False):
    """
        add line data + add style
        ARGS:
            ws, row_no, column_no, col_name="", val_1="", val_2="", str_diff="", other="", diff_flg=False
        RETRUN:
            ws
    """
    if str_diff is not None:
        # col_name
        ws.cell(row=row_no, column=column_no, value=wr_fmt(col_name))
        # diff
        ws.cell(row=row_no, column=column_no + 1, value=wr_fmt(str_diff))
        # value1: target_menu_1
        ws.cell(row=row_no, column=column_no + 2, value=wr_fmt(val_1))
        # value2: target_menu_2
        ws.cell(row=row_no, column=column_no + 3, value=wr_fmt(val_2))
        # detail
        ws.cell(row=row_no, column=column_no + 4, value=wr_fmt(other))

    # 等号対応: set data_type->str
    _target_cols = [col_name, str_diff, val_1, val_2, other]
    for _idx in range(0, len(_target_cols)):
        _c_no = column_no + _idx
        ws.cell(row=row_no, column=_c_no).number_format = openpyxl.styles.numbers.FORMAT_TEXT
        ws.cell(row=row_no, column=_c_no).data_type = 's'

    # add diff style
    if diff_flg is True:
        ws.cell(row=row_no, column=column_no + 1).font = Font(color='FF0000')
        ws.cell(row=row_no, column=column_no + 2).font = Font(color='FF0000')
        ws.cell(row=row_no, column=column_no + 3).font = Font(color='FF0000')

    if wrap_text_flg is True:
        ws.cell(row=row_no, column=column_no + 2).alignment = Alignment(wrapText=True)
        ws.cell(row=row_no, column=column_no + 3).alignment = Alignment(wrapText=True)
    return ws


# Excel:セル入力用の関連データ取得
def get_col_name_data(compare_data, row_no, target_host, col_name, compare_target_flg):
    """
        add line data + add style
        ARGS:
            compare_data, row_no, target_host, col_name, compare_target_flg
        RETRUN:
            row_no, val_1, val_2, val_diff_flg, str_diff,
    """
    row_no = row_no + 1
    # get value
    val_1 = compare_data.get(target_host, {}).get('target_data_1', {}).get(col_name)
    val_2 = compare_data.get(target_host, {}).get('target_data_2', {}).get(col_name)
    # get diff flg
    val_diff_flg = compare_data.get(target_host, {}).get('_data_diff_flg', {}).get(col_name)

    # diff_flg override: _data_diff_flg or _file_compare_execute_flg
    file_diff_flg = compare_data.get(target_host, {}).get('_file_compare_execute_flg', {}).get(col_name)
    val_diff_flg = val_diff_flg or file_diff_flg if file_diff_flg is not None else val_diff_flg
    val_diff_flg = val_diff_flg if isinstance(val_diff_flg, bool) else False

    if compare_target_flg is True:
        # convert compare diff flg
        if val_diff_flg is True:
            # 差分あり / ●
            str_diff = g.appmsg.get_api_message("MSG-60008")

        elif val_diff_flg is False:
            # 差分なし / ""
            str_diff = g.appmsg.get_api_message("MSG-60009")
        else:
            str_diff = None
            row_no = row_no - 1
    else:
        str_diff = ""

    return row_no, val_1, val_2, val_diff_flg, str_diff,


# Excel:シート作成 比較条件+比較結果一覧(summary)
def ws_index_create_table(wb, file_name, exec_time, config, compare_config, compare_diff_flg, tbl_start_str, tbl_end_str):
    """
        crate sheet compare parameter + summary
        ARGS:
            wb, file_name, exec_time, config, compare_config, compare_diff_flg, tbl_start_str, tbl_end_str
        RETRUN:
            ws
    """
    compare = compare_config.get("parameter").get("compare")
    base_date_1 = compare_config.get("parameter").get("base_date_1")
    base_date_2 = compare_config.get("parameter").get("base_date_2")
    host_list = compare_config.get("parameter").get("host")
    host_str = ",".join([str(_) for _ in host_list])

    detail_flg = _get_flg(compare_config, "detail_flg")
    target_menyu_name_1 = config.get("target_menus")[0]
    target_menyu_name_2 = config.get("target_menus")[1]
    target_host_list = config.get("target_host_list")

    # sheet name:比較条件_結果一覧
    target_sheet = wr_fmt(g.appmsg.get_api_message("MSG-60010"))
    ws = wb.active
    ws.title = target_sheet
    ws = wb[target_sheet]
    row_no = 1
    column_no = 1
    tmp_row_no = 0

    # 比較実行条件 値
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt(g.appmsg.get_api_message("MSG-60011")))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(g.appmsg.get_api_message("MSG-60012")))
    # 比較実行日時
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt(g.appmsg.get_api_message("MSG-60013")))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(exec_time))
    # 比較設定名
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt(g.appmsg.get_api_message("MSG-60014")))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(compare))

    # 詳細フラグ
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt("詳細設定フラグ"))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(detail_flg))

    # 比較対象メニュー1
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt("比較対象メニュー1"))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(target_menyu_name_1))

    # 比較対象メニュー2
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt("比較対象メニュー2"))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(target_menyu_name_2))

    # 基準日時1
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt(g.appmsg.get_api_message("MSG-60015")))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(base_date_1))
    # 基準日時2
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt(g.appmsg.get_api_message("MSG-60016")))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(base_date_2))
    # 対象ホスト
    tmp_row_no += 1
    ws.cell(row=tmp_row_no, column=column_no, value=wr_fmt(g.appmsg.get_api_message("MSG-60017")))
    ws.cell(row=tmp_row_no, column=column_no + 1, value=wr_fmt(host_str))

    # 等号対応: set data_type->str
    for _n in range(1, 10):
        ws.cell(row=_n, column=column_no).number_format = openpyxl.styles.numbers.FORMAT_TEXT
        ws.cell(row=_n, column=column_no + 1).data_type = 's'

    # add table: comapre execute parameters
    tbl_end_no = int("{}".format(tmp_row_no))
    table_name = "compare_parameter"
    table_style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ref_area = "{}{}:{}{}".format(tbl_start_str, 1, "B", tbl_end_no)
    ws = ws_add_table(ws, table_name, None, ref_area)

    # set data: compare summary
    # table header
    row_no = int("{}".format(tbl_end_no)) + 2
    tbl_start_no = int("{}".format(row_no))
    # 比較対象ホスト, 比較結果:対象シートにリンク
    ws.cell(row=row_no, column=column_no, value=wr_fmt(g.appmsg.get_api_message("MSG-60018")))
    ws.cell(row=row_no, column=column_no + 1, value=wr_fmt(g.appmsg.get_api_message("MSG-60019")))

    # table data
    row_no = row_no + 1
    for target_host in target_host_list:
        column_no = 1

        # create sheet: target_host
        target_sheet = wr_fmt(target_host)
        wb.create_sheet(target_sheet)

        # compare diff flg convert
        host_compare_diff_flg = compare_diff_flg.get(target_host)
        if host_compare_diff_flg is True:
            # 差分あり
            str_host_compare_diff = g.appmsg.get_api_message("MSG-60020")
            host_result_font = Font(color='FF0000', underline="single")
        else:
            # 差分なし
            str_host_compare_diff = g.appmsg.get_api_message("MSG-60021")
            host_result_font = Font(color='0000FF', underline="single")

        # set summary list table data
        ws.cell(row=row_no, column=column_no, value=wr_fmt(target_host))
        ws.cell(row=row_no, column=column_no + 1, value=wr_fmt(str_host_compare_diff))
        ws.cell(row=row_no, column=column_no).font = host_result_font
        ws.cell(row=row_no, column=column_no + 1).font = host_result_font

        # 等号対応: set data_type->str
        for _n in range(column_no, column_no + 2):
            ws.cell(row=row_no, column=_n).number_format = openpyxl.styles.numbers.FORMAT_TEXT
            ws.cell(row=row_no, column=_n).data_type = 's'

        # set link
        target_cell = "{}{}".format("A", row_no)
        target_link = "{}#'{}'!{}".format(file_name, target_sheet, "A1")
        ws[target_cell].hyperlink = target_link
        # ws[target_cell].value = '=HYPERLINK("' + target_link + '", "'+ str_host_compare_diff + '")'
        row_no = row_no + 1

        # set column width
        ws = ws_auto_adjusted_width(ws)

    # add table: compare summary
    tbl_end_no = int("{}".format(row_no)) - 1
    table_name = "host_link_table"
    table_style = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ref_area = "{}{}:{}{}".format(tbl_start_str, tbl_start_no, "B", tbl_end_no)
    ws = ws_add_table(ws, table_name, None, ref_area)

    return ws


# Excel:シート作成 比較結果シート(対象ホスト)
def ws_target_host_create_table(wb, file_name, exec_time, config, compare_data, compare_diff_flg, tbl_start_str, tbl_end_str):
    """
        crate sheet compare result :target host
        ARGS:
            wb, config, compare_data, compare_diff_flg, tbl_start_str, tbl_end_str
        RETRUN:
            ws
    """

    target_host_list = config.get("target_host_list")
    target_column_info = config.get("target_column_info")
    for target_host in target_host_list:
        # set data: host compare detail
        target_sheet = wr_fmt(target_host)
        ws = wb[target_sheet]

        # compare diff flg convert
        host_compare_diff_flg = compare_diff_flg.get(target_host)
        if host_compare_diff_flg is True:
            ws.sheet_properties.tabColor = 'FF0000'

        # set point
        row_no = 1
        column_no = 1
        column_no = 1

        # get target menu
        val_1 = "1:{}".format(config.get("target_menus")[0])
        val_2 = "2:{}".format(config.get("target_menus")[1])
        val_1 = "{}:{}".format(g.appmsg.get_api_message("MSG-60025"), config.get("target_menus")[0])[0:255]
        val_2 = "{}:{}".format(g.appmsg.get_api_message("MSG-60026"), config.get("target_menus")[1])[0:255]

        # set table start point
        tbl_start_no = int("{}".format(row_no))

        # set table header
        # メニュー名 差分 詳細
        col_name = g.appmsg.get_api_message("MSG-60022")
        str_diff = g.appmsg.get_api_message("MSG-60023")
        other = g.appmsg.get_api_message("MSG-60024")
        ws = ws_add_table_data_cells(ws, row_no, column_no, col_name, val_1, val_2, str_diff, other)

        # set view table data
        for target_column in target_column_info:
            col_name = target_column.get("col_name")
            compare_target_flg = target_column.get("compare_target_flg")
            column_class_1 = target_column.get("column_class_1")
            column_class_2 = target_column.get("column_class_2")
            wrap_text_flg = False
            if column_class_1 in ["MultiTextColumn"] or column_class_2 in ["MultiTextColumn"]:
                wrap_text_flg = True

            if col_name not in [g.appmsg.get_api_message("MSG-60022")]:
                row_no, val_1, val_2, val_diff_flg, str_diff = \
                    get_col_name_data(compare_data, row_no, target_host, col_name, compare_target_flg)
                # set target data
                ws = ws_add_table_data_cells(ws, row_no, column_no, col_name, val_1, val_2, str_diff, "", val_diff_flg, wrap_text_flg)

        # set column width
        ws = ws_auto_adjusted_width(ws)

        # add table: host compare detail
        tbl_end_no = int("{}".format(row_no))
        table_name = "target_host_" + rand_str(10)

        # add: dummy line target column 0
        if tbl_start_no == tbl_end_no:
            row_no += 1
            ws = ws_add_table_data_cells(ws, row_no, column_no)
            tbl_end_no = int("{}".format(row_no))

        ref_area = "{}{}:{}{}".format(tbl_start_str, tbl_start_no, tbl_end_str, tbl_end_no)
        ws = ws_add_table(ws, table_name, None, ref_area)

        row_no = row_no + 2
        ws.cell(row=row_no, column=column_no, value=wr_fmt("TOP"))
        ws.cell(row=row_no, column=column_no).font = Font(color='0000FF', underline="single")
        target_cell = "{}{}".format("A", row_no)
        target_sheet = wr_fmt(g.appmsg.get_api_message("MSG-60010"))
        target_link = "{}#'{}'!{}".format(file_name, target_sheet, "A1")
        ws[target_cell].hyperlink = target_link
    return ws


# mime判定不可ファイルのバイナリ判定
def no_mimetype_is_binary_chk(target_file_path, file_mimetype, encoding):
    """
        check binary file :no mimetype file
        ARGS:
            target_file_path
            file_mimetype
            encoding
        RETRUN:
        (ret, file_mimetype, encode)
    """
    import chardet
    ret = False
    encode = encoding
    # /storage配下のファイルアクセスを/tmp経由で行うモジュール
    file_read = storage_access.storage_read_bytes()

    if file_mimetype is None:
        # check encode -> check ASCII -08
        fd = file_read.read_bytes(target_file_path)
        encode = chardet.detect(fd)['encoding']
        if encode is None:
            ret = True
        else:
            tmp_code = [fdcode for fdcode in list(fd)]
            # check ASCII -08
            for n in range(0, 9):
                if n in tmp_code:
                    ret = True
                    break
    return ret, file_mimetype, encode,


# add filename lineno
def addline_msg(msg=''):
    import os
    import inspect
    info = inspect.getouterframes(inspect.currentframe())[1]
    msg_line = "{} ({}:{})".format(msg, os.path.basename(info.filename), info.lineno)
    return msg_line
